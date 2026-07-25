"""Local, auditable helpers for preparing university reimbursement materials.

The module deliberately keeps OCR, matching, and Excel rendering independent
from Flask. It can therefore be reused by a future desktop application without
requiring a server, external LLM, or database connection.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal, DecimalException, InvalidOperation, ROUND_HALF_UP
from math import gcd
from pathlib import Path
import re
import shutil
from typing import Callable, Iterable, TYPE_CHECKING
from zipfile import BadZipFile

from openpyxl import load_workbook

from core.inventory_line_extraction import (
    RecognizedLineItem,
    TextBox,
    extract_line_items_from_boxes,
)

if TYPE_CHECKING:
    from core.ocr import OCRResult


AMOUNT_BUCKETS = (
    (5_000, "0-50"),
    (10_000, "50-100"),
    (20_000, "100-200"),
    (50_000, "200-500"),
    (100_000, "500-1000"),
)

INBOUND_INVENTORY_HEADERS = (
    "入库日期", "产品名称", "规格型号", "单位", "入库数量",
    "单价(元)", "金额(元)", "供货单位", "经办人", "管理员",
)
OUTBOUND_INVENTORY_HEADERS = (
    "出库日期", "产品名称", "规格型号", "单位", "出库数量",
    "单价(元)", "金额(元)", "用途", "领用人", "管理员",
)
TOTAL_LABEL_COLUMN = 6
TOTAL_AMOUNT_COLUMN = 7
TOTAL_CAPITALIZED_LABEL_COLUMN = 8
TOTAL_CAPITALIZED_AMOUNT_COLUMN = 9
MAX_ALLOCATION_SEARCH_STATES = 200_000
MAX_SUBSET_CAPACITY_UNITS = 2_000_000


@dataclass(frozen=True)
class Receipt:
    """A source receipt after local text extraction or OCR."""

    receipt_id: str
    source_path: str
    invoice_date: date | None
    total_cents: int | None
    vendor_name: str = ""
    invoice_number: str = ""
    ocr_text: str = ""
    is_material: bool = True
    extraction_note: str = ""
    line_items: tuple[RecognizedLineItem, ...] = ()


@dataclass(frozen=True)
class CandidatePackage:
    """One non-final reimbursement candidate under a single amount cap."""

    capacity_cents: int
    total_cents: int
    receipt_ids: tuple[str, ...]

    @property
    def remaining_cents(self) -> int:
        return self.capacity_cents - self.total_cents


@dataclass(frozen=True)
class QuotaSlot:
    """A confirmed project balance that can hold one reimbursement package."""

    slot_id: str
    capacity_cents: int
    label: str = ""


@dataclass(frozen=True)
class QuotaPackage:
    """The selected receipts assigned to one confirmed quota slot."""

    slot_id: str
    capacity_cents: int
    total_cents: int
    receipt_ids: tuple[str, ...]
    label: str = ""

    @property
    def remaining_cents(self) -> int:
        return self.capacity_cents - self.total_cents


@dataclass(frozen=True)
class QuotaAllocation:
    """A globally consistent allocation across all confirmed quota slots."""

    packages: tuple[QuotaPackage, ...]
    unassigned_receipt_ids: tuple[str, ...]

    @property
    def allocated_total_cents(self) -> int:
        return sum(package.total_cents for package in self.packages)


@dataclass(frozen=True)
class InventoryLine:
    """A manually verified material line used by both inventory documents."""

    inventory_date: date
    name: str
    specification: str
    unit: str
    quantity: str | int | float | Decimal
    unit_price_cents: int
    amount_cents: int
    supplier_name: str = ""


def amount_bucket(amount_cents: int | None) -> str:
    """Return the agreed deterministic amount band for a tax-inclusive amount."""
    if amount_cents is None or amount_cents <= 0:
        return "invalid"
    for upper_bound, label in AMOUNT_BUCKETS:
        if amount_cents <= upper_bound:
            return label
    return "1000+"


def parse_amount_to_cents(value: str | int | float | Decimal) -> int:
    """Parse a tax-inclusive amount into integer cents without float rounding."""
    normalized = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    try:
        amount = Decimal(normalized).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid amount: {value!r}") from exc
    return int(amount * 100)


def parse_receipt_text(text: str, source_path: str = "") -> Receipt:
    """Extract the stable fields needed for local review from invoice text.

    This parser intentionally leaves item normalization to a review step. It
    never invents an item, amount, or date when those fields cannot be found.
    """
    compact = re.sub(r"\s+", "", text or "")
    invoice_number = _first_group(r"发票号(?:码)?[：:]?(\d{8,})", compact)
    if not invoice_number:
        invoice_number = _first_group(r"(\d{20})(?=20\d{2}[年\-/.])", compact)
    parsed_date = _parse_invoice_date(compact)
    vendor_name = _extract_vendor_name(compact)
    total_cents = _extract_tax_inclusive_total(compact)
    path = Path(source_path)
    receipt_id = path.stem or invoice_number or "receipt"

    return Receipt(
        receipt_id=receipt_id,
        source_path=str(source_path),
        invoice_date=parsed_date,
        total_cents=total_cents,
        vendor_name=vendor_name,
        invoice_number=invoice_number,
        ocr_text=text or "",
    )


def extract_pdf_text(source_path: str | Path) -> str:
    """Read selectable text from an electronic PDF without using OCR."""
    pdf_reader, pdf_errors = _pdf_reader_types()
    if pdf_reader is None:
        return ""

    try:
        reader = pdf_reader(str(source_path))
        return "\n".join(page.extract_text() or "" for page in reader.pages).strip()
    except (OSError, ValueError, *pdf_errors):
        return ""


def extract_pdf_text_boxes(source_path: str | Path) -> list[TextBox]:
    """Read positioned selectable PDF text for local table recognition."""
    pdf_reader, pdf_errors = _pdf_reader_types()
    if pdf_reader is None:
        return []

    boxes: list[TextBox] = []
    try:
        reader = pdf_reader(str(source_path))
        for page_index, page in enumerate(reader.pages):
            page_height = float(page.mediabox.height)

            def visitor(text, cm, tm, _font, _size):
                if text and text.strip():
                    x, y = _pdf_text_position(cm, tm, page_height)
                    boxes.append(TextBox(text.strip(), x, y + page_index * 10_000))

            page.extract_text(visitor_text=visitor)
    except (OSError, ValueError, TypeError, IndexError, *pdf_errors):
        return []
    return boxes


def _pdf_reader_types():
    """Return an available PDF reader without requiring both package names."""
    try:
        from pypdf import PdfReader
        from pypdf.errors import PdfReadError
    except ImportError:
        try:
            from PyPDF2 import PdfReader
            from PyPDF2.errors import PdfReadError
        except ImportError:
            return None, ()
    return PdfReader, (PdfReadError,)


def _pdf_text_position(cm, tm, page_height: float) -> tuple[float, float]:
    """Compose PDF page and text transforms into top-down page coordinates."""
    x = float(cm[4]) + float(tm[4]) * float(cm[0]) + float(tm[5]) * float(cm[2])
    pdf_y = float(cm[5]) + float(tm[4]) * float(cm[1]) + float(tm[5]) * float(cm[3])
    return x, page_height - pdf_y


def verified_pdf_line_items(
    source_path: str | Path, expected_total_cents: int | None
) -> tuple[RecognizedLineItem, ...]:
    """Return only PDF line items that reconcile to the confirmed total."""
    source = Path(source_path)
    if source.suffix.lower() != ".pdf" or expected_total_cents is None or expected_total_cents <= 0:
        return ()
    items = _line_items_from_boxes(extract_pdf_text_boxes(source), expected_total_cents)
    return items if _has_verified_items(items) else ()


def ingest_receipt(
    source_path: str | Path,
    provider_name: str | None = None,
    recognizer: Callable[[str], "OCRResult"] | None = None,
) -> Receipt:
    """Read a receipt locally and retain any reconcilable item-level details."""
    source = Path(source_path)
    text = extract_pdf_text(source) if source.suffix.lower() == ".pdf" else ""

    if text:
        receipt = parse_receipt_text(text, str(source))
        return replace(
            receipt,
            line_items=verified_pdf_line_items(source, receipt.total_cents),
        )

    result = _recognize_locally(source, provider_name, recognizer)
    if result is None:
        return Receipt(
            receipt_id=source.stem or "receipt",
            source_path=str(source),
            invoice_date=None,
            total_cents=None,
            extraction_note=(
                "未提取到真实票据文本。mock OCR 仅用于测试，"
                "请安装/启用 PaddleOCR 或手工填写。"
            ),
        )
    receipt = parse_receipt_text(result.text, str(source))
    ocr_items = _line_items_from_ocr_result(result, receipt.total_cents)
    return replace(
        receipt,
        line_items=ocr_items if _has_verified_items(ocr_items) else (),
        extraction_note=result.error or receipt.extraction_note,
    )


def _recognize_locally(
    source: Path,
    provider_name: str | None,
    recognizer: Callable[[str], "OCRResult"] | None = None,
):
    from core.ocr import OCRResult, get_ocr_provider, recognize_safely

    if recognizer is not None:
        try:
            result = recognizer(str(source))
        except Exception as exc:
            return OCRResult(
                provider=provider_name or "paddle",
                error=f"Local OCR recognition failed: {exc}",
            )
        if isinstance(result, OCRResult):
            result.provider = result.provider or provider_name or "paddle"
            return result
        return OCRResult(
            provider=provider_name or "paddle",
            error=f"Local OCR recognizer returned unsupported result type: {type(result).__name__}",
        )

    provider = get_ocr_provider(provider_name)
    if getattr(provider, "name", "") == "mock":
        return None
    return recognize_safely(provider, str(source))


def _line_items_from_ocr_result(result, expected_total_cents: int | None) -> tuple[RecognizedLineItem, ...]:
    if result is None or not isinstance(result.raw, dict):
        return ()
    from core.ocr.paddle_provider import extract_text_boxes

    return _line_items_from_boxes(extract_text_boxes(result.raw.get("result")), expected_total_cents)


def _line_items_from_boxes(boxes: list[TextBox], expected_total_cents: int | None) -> tuple[RecognizedLineItem, ...]:
    return tuple(extract_line_items_from_boxes(boxes, expected_total_cents))


def _has_verified_items(items: tuple[RecognizedLineItem, ...]) -> bool:
    return bool(items) and all(item.confidence == "verified" for item in items)


def render_receipt_markdown(receipt: Receipt) -> str:
    """Render a human-readable audit record; it is not part of the final package."""
    amount = _format_cents(receipt.total_cents)
    invoice_date = receipt.invoice_date.isoformat() if receipt.invoice_date else "待确认"
    return "\n".join(
        (
            "# 票据识别记录",
            "",
            f"- 来源文件: {Path(receipt.source_path).name or '待确认'}",
            f"- 发票日期: {invoice_date}",
            f"- 发票号: {receipt.invoice_number or '待确认'}",
            f"- 销售方: {receipt.vendor_name or '待确认'}",
            f"- 价税合计: {amount}",
            f"- 金额档位: {amount_bucket(receipt.total_cents)}",
            "- 材料票据: " + ("是" if receipt.is_material else "否"),
            f"- 解析说明: {receipt.extraction_note or '无'}",
            "",
            "## 原始识别文本",
            "",
            "```text",
            receipt.ocr_text.strip(),
            "```",
            "",
        )
    )


def build_readable_filename(receipt: Receipt, sequence: int) -> str:
    """Build the required YYMMDD_sequence_amount source-file name."""
    if receipt.invoice_date is None:
        raise ValueError("Invoice date must be confirmed before renaming a receipt")
    if receipt.total_cents is None or receipt.total_cents <= 0:
        raise ValueError("Tax-inclusive amount must be confirmed before renaming a receipt")
    if sequence < 1:
        raise ValueError("Sequence must start at 1")

    extension = Path(receipt.source_path).suffix.lower() or ".pdf"
    return f"{receipt.invoice_date:%y%m%d}_{sequence:02d}_{_format_cents(receipt.total_cents)}{extension}"


def build_inventory_line_drafts(receipts: Iterable[Receipt]) -> list[dict[str, object]]:
    """Build reviewable one-line inventory drafts from confirmed receipts.

    The OCR text is only used to suggest a material name.  Every monetary value
    comes from the human-confirmed receipt fields, and an unrecognised item name
    remains blank so it cannot silently become an export-ready inventory line.
    """
    drafts: list[dict[str, object]] = []
    for receipt in receipts:
        if receipt.invoice_date is None or receipt.total_cents is None or receipt.total_cents <= 0:
            continue
        base = {
            "inventory_date": receipt.invoice_date.isoformat(),
            "supplier_name": receipt.vendor_name,
            "_auto_generated": True,
            "_source_receipt_id": receipt.receipt_id,
        }
        if _has_verified_items(receipt.line_items):
            for item in receipt.line_items:
                amount_cents = item.amount_cents
                drafts.append(base | {
                    "name": item.name,
                    "specification": item.specification,
                    "unit": item.unit,
                    "quantity": item.quantity,
                    "unit_price": _format_cents(item.unit_price_cents) if item.unit_price_cents is not None else "",
                    "amount": _format_cents(amount_cents) if amount_cents is not None else "",
                    "recognition_status": "已校验",
                })
            continue
        amount = _format_cents(receipt.total_cents)
        drafts.append(base | {
            "name": "",
            "specification": "",
            "unit": "",
            "quantity": "",
            "unit_price": "",
            "amount": amount,
            "recognition_status": "待人工填写",
        })
    return drafts


def copy_renamed_receipts(receipts: Iterable[Receipt], destination_dir: str | Path) -> list[Path]:
    """Copy original receipts into one package while preserving their extensions."""
    destination = Path(destination_dir)
    destination.mkdir(parents=True, exist_ok=True)
    copied_paths: list[Path] = []

    for sequence, receipt in enumerate(receipts, start=1):
        source = Path(receipt.source_path)
        if not source.is_file():
            raise FileNotFoundError(f"Receipt source file does not exist: {source}")
        target = destination / build_readable_filename(receipt, sequence)
        shutil.copy2(source, target)
        copied_paths.append(target)

    return copied_paths


def find_candidate_packages(
    receipts: Iterable[Receipt], capacity_cents: int, limit: int = 10
) -> list[CandidatePackage]:
    """Return high-utilization, non-overflowing candidate receipt combinations.

    Candidates can overlap because a user chooses one candidate before creating
    the next package. Within each candidate, a receipt always appears once.
    """
    if capacity_cents <= 0:
        raise ValueError("Capacity must be a positive integer number of cents")
    if limit <= 0:
        return []

    eligible = [
        receipt
        for receipt in receipts
        if receipt.is_material
        and receipt.total_cents is not None
        and 0 < receipt.total_cents <= capacity_cents
    ]
    states: dict[int, tuple[int, ...]] = {0: ()}

    for index, receipt in enumerate(eligible):
        for subtotal, indices in list(states.items()):
            candidate_total = subtotal + receipt.total_cents
            if candidate_total > capacity_cents:
                continue
            candidate_indices = indices + (index,)
            existing = states.get(candidate_total)
            if existing is None or len(candidate_indices) < len(existing):
                states[candidate_total] = candidate_indices

    packages = [
        CandidatePackage(
            capacity_cents=capacity_cents,
            total_cents=subtotal,
            receipt_ids=tuple(eligible[index].receipt_id for index in indices),
        )
        for subtotal, indices in states.items()
        if subtotal > 0
    ]
    packages.sort(key=lambda package: (-package.total_cents, len(package.receipt_ids), package.receipt_ids))
    return packages[:limit]


def allocate_receipts_to_quota_slots(
    receipts: Iterable[Receipt], slots: Iterable[QuotaSlot]
) -> QuotaAllocation:
    """Assign material receipts to confirmed balances without splitting receipts.

    A bounded exact search maximizes manageable batches. Fast subset-fill and
    best-fit candidates provide strong deterministic results before the search,
    so larger batches remain responsive instead of falling back solely because
    they crossed an arbitrary receipt-count threshold.
    """
    receipt_list = list(receipts)
    slot_list = list(slots)
    if len({slot.slot_id for slot in slot_list}) != len(slot_list):
        raise ValueError("Quota slot IDs must be unique")
    if any(slot.capacity_cents <= 0 for slot in slot_list):
        raise ValueError("Quota slot capacities must be positive")

    eligible = [
        receipt
        for receipt in receipt_list
        if receipt.is_material and receipt.total_cents is not None and receipt.total_cents > 0
    ]
    eligible.sort(key=lambda receipt: (-receipt.total_cents, receipt.receipt_id))

    assigned_indices = _allocate_exact(eligible, slot_list)

    packages: list[QuotaPackage] = []
    assigned_receipt_ids: set[str] = set()
    for slot_index, slot in enumerate(slot_list):
        indices = assigned_indices[slot_index]
        slot_receipts = [eligible[index] for index in indices]
        receipt_ids = tuple(receipt.receipt_id for receipt in slot_receipts)
        assigned_receipt_ids.update(receipt_ids)
        packages.append(
            QuotaPackage(
                slot_id=slot.slot_id,
                capacity_cents=slot.capacity_cents,
                total_cents=sum(receipt.total_cents or 0 for receipt in slot_receipts),
                receipt_ids=receipt_ids,
                label=slot.label,
            )
        )

    unassigned = tuple(
        receipt.receipt_id for receipt in receipt_list if receipt.receipt_id not in assigned_receipt_ids
    )
    return QuotaAllocation(packages=tuple(packages), unassigned_receipt_ids=unassigned)


def _allocate_exact(receipts: list[Receipt], slots: list[QuotaSlot]) -> list[list[int]]:
    suffix_totals = [0] * (len(receipts) + 1)
    for index in range(len(receipts) - 1, -1, -1):
        suffix_totals[index] = suffix_totals[index + 1] + (receipts[index].total_cents or 0)

    remaining = [slot.capacity_cents for slot in slots]
    current = [[] for _ in slots]
    best_assignment = _allocate_best_fit(receipts, slots)
    best_total = _assignment_total(receipts, best_assignment)
    maximum_total = min(suffix_totals[0], sum(remaining))

    for slot_order in _allocation_slot_orders(slots):
        candidate = _allocate_by_subset_fill(receipts, slots, slot_order)
        candidate_total = _assignment_total(receipts, candidate)
        if candidate_total > best_total:
            best_total = candidate_total
            best_assignment = candidate
        if best_total == maximum_total:
            return best_assignment

    visited: set[tuple[int, tuple[int, ...]]] = set()

    def search(receipt_index: int, assigned_total: int) -> None:
        nonlocal best_total, best_assignment
        if best_total == maximum_total:
            return
        if len(visited) >= MAX_ALLOCATION_SEARCH_STATES:
            return
        possible_total = assigned_total + min(
            suffix_totals[receipt_index], sum(remaining)
        )
        if possible_total <= best_total:
            return
        if receipt_index == len(receipts):
            if assigned_total > best_total:
                best_total = assigned_total
                best_assignment = [items[:] for items in current]
            return

        state_key = (receipt_index, tuple(sorted(remaining, reverse=True)))
        if state_key in visited:
            return
        visited.add(state_key)

        receipt = receipts[receipt_index]
        amount = receipt.total_cents or 0
        seen_remaining: set[int] = set()
        fitting_slots = sorted(
            (
                (slot_remaining - amount, slot_index, slot_remaining)
                for slot_index, slot_remaining in enumerate(remaining)
                if slot_remaining >= amount
            )
        )
        for _, slot_index, slot_remaining in fitting_slots:
            if slot_remaining < amount or slot_remaining in seen_remaining:
                continue
            seen_remaining.add(slot_remaining)
            remaining[slot_index] -= amount
            current[slot_index].append(receipt_index)
            search(receipt_index + 1, assigned_total + amount)
            current[slot_index].pop()
            remaining[slot_index] += amount

        search(receipt_index + 1, assigned_total)

    search(0, 0)
    return best_assignment


def _assignment_total(receipts: list[Receipt], assignment: list[list[int]]) -> int:
    return sum(
        receipts[index].total_cents or 0
        for slot_indices in assignment
        for index in slot_indices
    )


def _allocation_slot_orders(slots: list[QuotaSlot]) -> tuple[tuple[int, ...], ...]:
    original = tuple(range(len(slots)))
    orders = (
        original,
        tuple(sorted(original, key=lambda index: (slots[index].capacity_cents, index))),
        tuple(sorted(original, key=lambda index: (-slots[index].capacity_cents, index))),
    )
    return tuple(dict.fromkeys(orders))


def _allocate_by_subset_fill(
    receipts: list[Receipt], slots: list[QuotaSlot], slot_order: tuple[int, ...]
) -> list[list[int]]:
    available = list(range(len(receipts)))
    assignment = [[] for _ in slots]
    for slot_index in slot_order:
        selected = _best_subset_for_capacity(receipts, available, slots[slot_index].capacity_cents)
        assignment[slot_index] = selected
        selected_set = set(selected)
        available = [index for index in available if index not in selected_set]
    return assignment


def _best_subset_for_capacity(
    receipts: list[Receipt], available: list[int], capacity_cents: int
) -> list[int]:
    candidates = [
        index
        for index in available
        if 0 < (receipts[index].total_cents or 0) <= capacity_cents
    ]
    if not candidates:
        return []

    divisor = 0
    for index in candidates:
        divisor = gcd(divisor, receipts[index].total_cents or 0)
    divisor = max(divisor, 1)
    scaled_capacity = capacity_cents // divisor
    if scaled_capacity > MAX_SUBSET_CAPACITY_UNITS:
        return _best_fit_subset(receipts, candidates, capacity_cents)

    mask = (1 << (scaled_capacity + 1)) - 1
    reachable = 1
    history = [reachable]
    scaled_amounts: list[int] = []
    for index in candidates:
        amount = (receipts[index].total_cents or 0) // divisor
        scaled_amounts.append(amount)
        reachable = (reachable | (reachable << amount)) & mask
        history.append(reachable)

    subtotal = reachable.bit_length() - 1
    selected: list[int] = []
    for position in range(len(candidates) - 1, -1, -1):
        previous = history[position]
        if (previous >> subtotal) & 1:
            continue
        amount = scaled_amounts[position]
        selected.append(candidates[position])
        subtotal -= amount
    selected.reverse()
    return selected


def _best_fit_subset(
    receipts: list[Receipt], candidates: list[int], capacity_cents: int
) -> list[int]:
    remaining = capacity_cents
    selected: list[int] = []
    for index in candidates:
        amount = receipts[index].total_cents or 0
        if amount <= remaining:
            selected.append(index)
            remaining -= amount
    return selected


def _allocate_best_fit(receipts: list[Receipt], slots: list[QuotaSlot]) -> list[list[int]]:
    remaining = [slot.capacity_cents for slot in slots]
    assigned = [[] for _ in slots]
    for receipt_index, receipt in enumerate(receipts):
        amount = receipt.total_cents or 0
        fitting_slots = [
            (remaining[slot_index] - amount, slot_index)
            for slot_index in range(len(slots))
            if remaining[slot_index] >= amount
        ]
        if not fitting_slots:
            continue
        _, slot_index = min(fitting_slots)
        remaining[slot_index] -= amount
        assigned[slot_index].append(receipt_index)
    return assigned


def validate_inventory_lines(lines: Iterable[InventoryLine], expected_total_cents: int | None = None) -> int:
    """Validate material rows before any official-looking document is created."""
    line_list = list(lines)
    if not line_list:
        raise ValueError("At least one material line is required")

    for line in line_list:
        if not line.inventory_date:
            raise ValueError("Each material line requires an inventory date")
        if not line.name.strip() or not line.unit.strip():
            raise ValueError("Each material line requires a name and unit")
        if line.amount_cents <= 0 or line.unit_price_cents < 0:
            raise ValueError("Material amounts must be non-negative and line totals must be positive")
        try:
            quantity = Decimal(str(line.quantity))
        except DecimalException as exc:
            raise ValueError(f"Invalid material quantity: {line.quantity!r}") from exc
        if not quantity.is_finite() or quantity <= 0:
            raise ValueError(f"Invalid material quantity: {line.quantity!r}")
        try:
            calculated_cents = int(
                (quantity * Decimal(line.unit_price_cents)).quantize(
                    Decimal("1"), rounding=ROUND_HALF_UP
                )
            )
        except DecimalException as exc:
            raise ValueError(f"Invalid material quantity: {line.quantity!r}") from exc
        if abs(calculated_cents - line.amount_cents) > 1:
            raise ValueError(
                f"Material line {line.name!r} has quantity x unit-price total "
                f"{_format_cents(calculated_cents)}, not {_format_cents(line.amount_cents)}"
            )

    total_cents = sum(line.amount_cents for line in line_list)
    if expected_total_cents is not None and total_cents != expected_total_cents:
        raise ValueError(
            f"Material lines total {_format_cents(total_cents)} does not match "
            f"expected {_format_cents(expected_total_cents)}"
        )
    return total_cents


def validate_package_receipts(receipts: Iterable[Receipt]) -> int:
    """Return the tax-inclusive total for source receipts used by one package."""
    receipt_list = list(receipts)
    if not receipt_list:
        raise ValueError("A reimbursement package requires at least one source receipt")

    receipt_ids = [receipt.receipt_id for receipt in receipt_list]
    if len(receipt_ids) != len(set(receipt_ids)):
        raise ValueError("A reimbursement package cannot use the same receipt twice")
    if any(not receipt.is_material for receipt in receipt_list):
        raise ValueError("Non-material receipts cannot be used in inventory documents")
    if any(receipt.total_cents is None or receipt.total_cents <= 0 for receipt in receipt_list):
        raise ValueError("Every source receipt requires a confirmed tax-inclusive amount")
    if any(receipt.invoice_date is None for receipt in receipt_list):
        raise ValueError("Every source receipt requires a confirmed invoice date")
    return sum(receipt.total_cents for receipt in receipt_list if receipt.total_cents is not None)


def render_inventory_documents(
    template_dir: str | Path,
    output_dir: str | Path,
    lines: Iterable[InventoryLine],
    expected_total_cents: int | None = None,
) -> dict[str, Path]:
    """Render one paired inbound/outbound workbook from supplied official templates."""
    line_list = list(lines)
    validate_inventory_lines(line_list, expected_total_cents)
    templates = Path(template_dir)
    inbound_template = _find_template(templates, "入库单")
    outbound_template = _find_template(templates, "出库单")
    if inbound_template == outbound_template:
        raise ValueError("入库单和出库单模板不能是同一文件")

    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    inbound_output = output / "入库单.xlsx"
    outbound_output = output / "出库单.xlsx"

    _render_inventory_form(inbound_template, inbound_output, line_list, is_inbound=True)
    _render_inventory_form(outbound_template, outbound_output, line_list, is_inbound=False)
    return {"inbound": inbound_output, "outbound": outbound_output}


def _first_group(pattern: str, text: str) -> str:
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def _parse_invoice_date(text: str) -> date | None:
    match = re.search(r"(20\d{2})[年\-/.](\d{1,2})[月\-/.](\d{1,2})日?", text)
    if not match:
        return None
    try:
        return date(*(int(value) for value in match.groups()))
    except ValueError:
        return None


def _extract_vendor_name(text: str) -> str:
    patterns = (
        r"销售方信息名称[：:]?(.*?)(?:统一社会信用代码|纳税人识别号|项目名称)",
        r"销售方名称[：:]?(.*?)(?:统一社会信用代码|纳税人识别号|项目名称)",
        r"销售方信息名称[：:]?(.*?)(?:价税合计|合计)",
        r"销售方名称[：:]?(.*?)(?:价税合计|合计)",
        r"商家[：:]?(.*?)(?:日期|金额|票据号|发票号)",
    )
    for pattern in patterns:
        value = _first_group(pattern, text)
        if value:
            return value
    fallback = _first_group(
        r"([\u4e00-\u9fff（）()]+(?:有限责任公司|有限公司|商行(?:（个体工商户）)?|经营部))"
        r"(?:统一社会信用代码/纳税人识别号[:：]?)?(?=[A-Z0-9]{18})",
        text,
    )
    if fallback:
        return fallback
    return ""


def _extract_tax_inclusive_total(text: str) -> int | None:
    explicit_patterns = (
        r"价税合计[（(]小写[）)][：:]?[¥￥]?(\d{1,12}(?:\.\d{1,2})?)(?!\d)",
        r"价税合计[：:]?[¥￥](\d{1,12}(?:\.\d{1,2})?)(?!\d)",
    )
    for pattern in explicit_patterns:
        explicit_totals = re.findall(pattern, text)
        if explicit_totals:
            return parse_amount_to_cents(explicit_totals[-1])

    currency_amounts = re.findall(r"[¥￥](\d+(?:\.\d{1,2})?)", text)
    if currency_amounts:
        return max(parse_amount_to_cents(amount) for amount in currency_amounts)

    labelled_amounts = re.findall(r"(?:价税合计|金额)[^0-9]{0,12}(\d+(?:\.\d{1,2})?)", text)
    if labelled_amounts:
        return parse_amount_to_cents(labelled_amounts[-1])
    return None


def _format_cents(amount_cents: int | None) -> str:
    if amount_cents is None:
        return "待确认"
    return f"{Decimal(amount_cents) / Decimal(100):.2f}"


def _find_template(template_dir: Path, keyword: str) -> Path:
    matches = find_inventory_templates(template_dir, keyword)
    if not matches:
        raise FileNotFoundError(f"No {keyword} template found in {template_dir}")
    if len(matches) != 1:
        raise ValueError(f"{keyword}模板必须恰好一个，当前找到 {len(matches)} 个")
    try:
        inspect_inventory_template(matches[0], is_inbound=keyword == "入库单")
    except (BadZipFile, OSError, ValueError, KeyError) as exc:
        raise ValueError(f"{keyword}模板不可用：{matches[0].name}（{exc}）") from exc
    return matches[0]


def find_inventory_templates(template_dir: str | Path, keyword: str) -> tuple[Path, ...]:
    """Return Excel templates whose names carry the required inventory-form keyword."""
    directory = Path(template_dir)
    return tuple(sorted(path for path in directory.glob("*.xlsx") if keyword in path.name))


def inspect_inventory_template(
    template_path: str | Path, is_inbound: bool | None = None
) -> tuple[int, int]:
    """Locate the required header and total rows in an inventory workbook template."""
    workbook = load_workbook(template_path)
    try:
        sheet = workbook.active
        expected_headers = (
            INBOUND_INVENTORY_HEADERS if is_inbound else OUTBOUND_INVENTORY_HEADERS
            if is_inbound is not None else None
        )
        header_row = _find_header_row(sheet, expected_headers)
        return header_row, _find_total_row(sheet, header_row)
    finally:
        workbook.close()


def _render_inventory_form(
    template_path: Path, output_path: Path, lines: list[InventoryLine], is_inbound: bool
) -> None:
    workbook = load_workbook(template_path)
    sheet = workbook.active
    header_row = _find_header_row(
        sheet, INBOUND_INVENTORY_HEADERS if is_inbound else OUTBOUND_INVENTORY_HEADERS
    )
    data_start_row = header_row + 1
    total_row = _find_total_row(sheet, header_row)
    template_total_row = total_row
    existing_line_rows = total_row - data_start_row
    extra_rows = max(0, len(lines) - existing_line_rows)
    original_total_formula = sheet.cell(total_row, TOTAL_CAPITALIZED_AMOUNT_COLUMN).value

    if extra_rows:
        _insert_styled_rows(sheet, total_row, extra_rows, data_start_row)
        total_row += extra_rows

    for row_number in range(data_start_row, total_row):
        for column in range(1, 11):
            sheet.cell(row_number, column).value = None

    for offset, line in enumerate(lines):
        row_number = data_start_row + offset
        sheet.cell(row_number, 1).value = line.inventory_date
        sheet.cell(row_number, 1).number_format = "yyyy-mm-dd"
        _set_excel_text(sheet.cell(row_number, 2), line.name)
        _set_excel_text(sheet.cell(row_number, 3), line.specification)
        _set_excel_text(sheet.cell(row_number, 4), line.unit)
        sheet.cell(row_number, 5).value = _to_excel_number(line.quantity)
        sheet.cell(row_number, 6).value = line.unit_price_cents / 100
        sheet.cell(row_number, 7).value = line.amount_cents / 100
        sheet.cell(row_number, 6).number_format = "0.00"
        sheet.cell(row_number, 7).number_format = "0.00"
        if is_inbound:
            _set_excel_text(sheet.cell(row_number, 8), line.supplier_name)

    last_line_row = data_start_row + len(lines) - 1
    sheet.cell(total_row, TOTAL_LABEL_COLUMN).value = "合计："
    sheet.cell(total_row, TOTAL_AMOUNT_COLUMN).value = f"=SUM(G{data_start_row}:G{last_line_row})"
    sheet.cell(total_row, TOTAL_AMOUNT_COLUMN).number_format = "0.00"
    sheet.cell(total_row, TOTAL_CAPITALIZED_LABEL_COLUMN).value = "大写："
    sheet.cell(total_row, TOTAL_CAPITALIZED_AMOUNT_COLUMN).value = _update_total_formula(
        original_total_formula, template_total_row, total_row
    )

    sheet.print_title_rows = f"1:{header_row}"
    sheet.print_area = f"A1:J{sheet.max_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    workbook.save(output_path)


def _find_header_row(sheet, expected_headers: tuple[str, ...] | None = None) -> int:
    allowed_headers = (expected_headers,) if expected_headers is not None else (
        INBOUND_INVENTORY_HEADERS,
        OUTBOUND_INVENTORY_HEADERS,
    )
    for row in sheet.iter_rows(max_col=10):
        headers = tuple(str(cell.value or "").strip() for cell in row[:10])
        if any(_headers_match_template_variant(headers, expected) for expected in allowed_headers):
            return row[0].row
    raise ValueError("Template is missing the required 10-column header row")


def _headers_match_template_variant(
    headers: tuple[str, ...], expected_headers: tuple[str, ...]
) -> bool:
    """Accept the official short ``规格`` header without relaxing column order."""
    if len(headers) != len(expected_headers):
        return False
    return all(
        actual == expected
        or (index == 2 and expected == "规格型号" and actual == "规格")
        for index, (actual, expected) in enumerate(zip(headers, expected_headers))
    )


def _find_total_row(sheet, header_row: int) -> int:
    for row in sheet.iter_rows(min_row=header_row + 1):
        total_row = row[0].row
        if str(sheet.cell(total_row, TOTAL_LABEL_COLUMN).value or "").strip() == "合计：":
            _validate_total_row(sheet, header_row, total_row)
            return total_row
        if any(str(cell.value or "").strip() == "合计：" for cell in row):
            raise ValueError("Template 合计： label must be in column F")
    raise ValueError("Template is missing a 合计： row after the header")


def _validate_total_row(sheet, header_row: int, total_row: int) -> None:
    expected_amount_formula = f"=SUM(G{header_row + 1}:G{total_row - 1})"
    if sheet.cell(total_row, TOTAL_AMOUNT_COLUMN).value != expected_amount_formula:
        raise ValueError("Template total amount formula must sum the amount column data rows")
    if str(sheet.cell(total_row, TOTAL_CAPITALIZED_LABEL_COLUMN).value or "").strip() != "大写：":
        raise ValueError("Template total uppercase label must be in column H")
    capitalized_total_formula = sheet.cell(total_row, TOTAL_CAPITALIZED_AMOUNT_COLUMN).value
    total_cell_reference = rf"(?<![A-Z0-9_])\$?G\$?{total_row}(?!\d)"
    if not (
        isinstance(capitalized_total_formula, str)
        and capitalized_total_formula.startswith("=")
        and re.search(total_cell_reference, capitalized_total_formula, flags=re.IGNORECASE)
    ):
        raise ValueError("Template capitalized total formula must reference the total amount cell")


def _insert_styled_rows(sheet, insertion_row: int, amount: int, style_source_row: int) -> None:
    merges_to_shift = [
        (item.min_col, item.min_row, item.max_col, item.max_row)
        for item in sheet.merged_cells.ranges
        if item.min_row >= insertion_row
    ]
    for min_col, min_row, max_col, max_row in merges_to_shift:
        sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

    sheet.insert_rows(insertion_row, amount)

    source_row = max(style_source_row, insertion_row - 1)
    for row_number in range(insertion_row, insertion_row + amount):
        for column in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(source_row, column)
            target_cell = sheet.cell(row_number, column)
            target_cell._style = copy(source_cell._style)
            if source_cell.has_style:
                target_cell.number_format = source_cell.number_format
        sheet.row_dimensions[row_number].height = sheet.row_dimensions[source_row].height

    for min_col, min_row, max_col, max_row in merges_to_shift:
        sheet.merge_cells(
            start_row=min_row + amount,
            start_column=min_col,
            end_row=max_row + amount,
            end_column=max_col,
        )


def _update_total_formula(formula: object, source_total_row: int, total_row: int) -> str:
    if isinstance(formula, str) and formula.startswith("="):
        total_reference = re.compile(
            rf"(?<![A-Z0-9_])(\$?G)(\$?){source_total_row}(?!\d)", re.IGNORECASE
        )
        return total_reference.sub(
            lambda match: f"{match.group(1)}{match.group(2)}{total_row}", formula
        )
    return f'=TEXT(G{total_row},"[DBNUM2]")&"元整"'


def _to_excel_number(value: str | int | float | Decimal) -> int | float | str:
    try:
        decimal_value = Decimal(str(value))
    except InvalidOperation:
        return str(value)
    if decimal_value == decimal_value.to_integral():
        return int(decimal_value)
    return float(decimal_value)


def _set_excel_text(cell, value: object) -> None:
    """Store user-controlled workbook fields as literal text, never formulas."""
    cell.value = str(value)
    cell.data_type = "s"
