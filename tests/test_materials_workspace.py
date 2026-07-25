from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
import json
import unittest
from unittest.mock import patch
from zipfile import BadZipFile, ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from core.materials_assistant import QuotaSlot, Receipt
from core.inventory_line_extraction import RecognizedLineItem
from core.materials_workspace import (
    ExportResult,
    assign_receipt_to_slot,
    package_preflight,
    create_workspace,
    export_workspace_backup,
    export_quota_package,
    load_workspace,
    migrate_legacy_workspace,
    restore_workspace_backup,
    save_workspace,
    validate_template_directory,
    workspace_quotas,
    workspace_receipts,
)


class MaterialsWorkspaceTests(unittest.TestCase):
    def test_workspace_round_trip_preserves_recognized_line_items(self):
        with TemporaryDirectory() as directory:
            workspace, state = create_workspace(Path(directory), "project", "templates")
            state["receipts"] = [{
                "receipt_id": "receipt-1",
                "source_path": "receipt.pdf",
                "invoice_date": "2026-06-20",
                "total_cents": 96_600,
                "line_items": [{
                    "name": "RoboMaster C620无刷电机调速器",
                    "specification": "C620",
                    "unit": "件",
                    "quantity": "6",
                    "unit_price_cents": 16_100,
                    "amount_cents": 96_600,
                    "confidence": "verified",
                }],
            }]
            save_workspace(workspace, state)

            restored = workspace_receipts(load_workspace(workspace))[0]

        self.assertEqual(
            restored.line_items,
            (RecognizedLineItem("RoboMaster C620无刷电机调速器", "C620", "件", "6", 16_100, 96_600, "verified"),),
        )
    def test_assign_receipt_to_slot_reassigns_without_duplicates_and_can_unassign(self):
        with TemporaryDirectory() as directory:
            state = self._valid_package_state(Path(directory))
            state["quotas"].append({
                "slot_id": "quota-2", "capacity_cents": 20_000, "receipt_ids": [],
            })

            assign_receipt_to_slot(state, "receipt-1", "quota-2")
            assign_receipt_to_slot(state, "receipt-1", "quota-2")

            self.assertEqual(state["quotas"][0]["receipt_ids"], [])
            self.assertEqual(state["quotas"][1]["receipt_ids"], ["receipt-1"])
            assign_receipt_to_slot(state, "receipt-1", None)
            self.assertEqual(state["quotas"][1]["receipt_ids"], [])

    def test_assign_receipt_to_slot_rejects_unusable_or_over_quota_receipts(self):
        for field, value in (("confirmed", False), ("is_material", False), ("total_cents", 20_000)):
            with self.subTest(field=field), TemporaryDirectory() as directory:
                state = self._valid_package_state(Path(directory))
                state["quotas"].append({
                    "slot_id": "quota-2", "capacity_cents": 10_000, "receipt_ids": [],
                })
                state["receipts"][0][field] = value

                with self.assertRaises(ValueError):
                    assign_receipt_to_slot(state, "receipt-1", "quota-2")

                self.assertEqual(state["quotas"][0]["receipt_ids"], ["receipt-1"])
                self.assertEqual(state["quotas"][1]["receipt_ids"], [])
    def test_template_validation_requires_both_inventory_templates(self):
        with TemporaryDirectory() as directory:
            templates = Path(directory) / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx")

            validation = validate_template_directory(templates)

            self.assertFalse(validation.valid)
            self.assertIsNotNone(validation.inbound_path)
            self.assertIsNone(validation.outbound_path)
            self.assertIn("出库单模板", validation.errors[0])

    def test_template_validation_accepts_a_usable_inbound_and_outbound_pair(self):
        with TemporaryDirectory() as directory:
            templates = Path(directory) / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx")
            self._write_template(templates / "出库单_模板.xlsx")

            validation = validate_template_directory(templates)

            self.assertTrue(validation.valid)
            self.assertEqual(validation.inbound_path, (templates / "入库单_模板.xlsx").resolve())
            self.assertEqual(validation.outbound_path, (templates / "出库单_模板.xlsx").resolve())
            self.assertEqual(validation.errors, ())

    def test_template_validation_accepts_official_short_specification_headers(self):
        with TemporaryDirectory() as directory:
            templates = Path(directory) / "templates"
            templates.mkdir()
            self._write_template(
                templates / "入库单_模板.xlsx",
                headers=[
                    "入库日期", "产品名称", "规格", "单位", "入库数量",
                    "单价(元)", "金额(元)", "供货单位", "经办人", "管理员",
                ],
            )
            self._write_template(
                templates / "出库单_模板.xlsx",
                headers=[
                    "出库日期", "产品名称", "规格", "单位", "出库数量",
                    "单价(元)", "金额(元)", "用途", "领用人", "管理员",
                ],
            )

            validation = validate_template_directory(templates)

            self.assertTrue(validation.valid)
            self.assertEqual(validation.errors, ())

    def test_template_validation_rejects_missing_or_misaligned_columns(self):
        inbound_headers = [
            "入库日期", "产品名称", "规格型号", "单位", "入库数量",
            "单价(元)", "金额(元)", "供货单位", "经办人", "管理员",
        ]
        invalid_headers = (
            inbound_headers[:-1],
            [inbound_headers[0], inbound_headers[2], inbound_headers[1], *inbound_headers[3:]],
        )
        for headers in invalid_headers:
            with self.subTest(headers=headers), TemporaryDirectory() as directory:
                templates = Path(directory) / "templates"
                templates.mkdir()
                self._write_template(templates / "入库单_模板.xlsx", headers=headers)
                self._write_template(templates / "出库单_模板.xlsx")

                validation = validate_template_directory(templates)

                self.assertFalse(validation.valid)
                self.assertTrue(any("入库单模板不可用" in error for error in validation.errors))

    def test_template_validation_rejects_total_row_before_header(self):
        with TemporaryDirectory() as directory:
            templates = Path(directory) / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx", total_row=3)
            self._write_template(templates / "出库单_模板.xlsx")

            validation = validate_template_directory(templates)

            self.assertFalse(validation.valid)
            self.assertTrue(any("入库单模板不可用" in error for error in validation.errors))

    def test_template_validation_rejects_misplaced_total_label_or_missing_total_formulas(self):
        invalid_totals = (
            {"total_label_column": 5},
            {"amount_formula": False},
            {"capitalized_total_formula": False},
        )
        for template_options in invalid_totals:
            with self.subTest(template_options=template_options), TemporaryDirectory() as directory:
                templates = Path(directory) / "templates"
                templates.mkdir()
                self._write_template(templates / "入库单_模板.xlsx", **template_options)
                self._write_template(templates / "出库单_模板.xlsx")

                validation = validate_template_directory(templates)

                self.assertFalse(validation.valid)
                self.assertTrue(any("入库单模板不可用" in error for error in validation.errors))

    def test_template_validation_rejects_an_unreadable_workbook(self):
        with TemporaryDirectory() as directory:
            templates = Path(directory) / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx")
            (templates / "出库单_模板.xlsx").write_bytes(b"not an xlsx workbook")

            validation = validate_template_directory(templates)

            self.assertFalse(validation.valid)
            self.assertTrue(any("出库单模板不可用" in error for error in validation.errors))

    def test_template_validation_rejects_single_workbook_used_for_both_forms(self):
        with TemporaryDirectory() as directory:
            templates = Path(directory) / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单出库单_模板.xlsx")

            validation = validate_template_directory(templates)

            self.assertFalse(validation.valid)
            self.assertTrue(any("出库单模板不可用" in error for error in validation.errors))

    def test_package_preflight_rejects_unconfirmed_receipt(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["receipts"][0]["confirmed"] = False

            errors = package_preflight(state, "quota-1")

            self.assertIn("票据 receipt-1 尚未人工确认", errors)

    def test_package_preflight_rejects_material_total_that_does_not_match_receipts(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["lines_by_slot"]["quota-1"][0]["quantity"] = "1"
            state["lines_by_slot"]["quota-1"][0]["unit_price"] = "99.99"
            state["lines_by_slot"]["quota-1"][0]["amount"] = "99.99"

            errors = package_preflight(state, "quota-1")

            self.assertTrue(any("物品明细合计与票据价税合计不一致" in error for error in errors))

    def test_package_preflight_rejects_receipts_over_quota(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["quotas"][0]["capacity_cents"] = 9_999

            errors = package_preflight(state, "quota-1")

            self.assertTrue(any("票据合计超过额度" in error for error in errors))

    def test_package_preflight_rejects_non_material_receipt(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["receipts"][0]["is_material"] = False

            errors = package_preflight(state, "quota-1")

            self.assertIn("票据 receipt-1 不是材料票据", errors)

    def test_package_preflight_rejects_missing_inventory_lines(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["lines_by_slot"]["quota-1"] = []

            errors = package_preflight(state, "quota-1")

            self.assertIn("额度 quota-1 无物品明细", errors)

    def test_package_preflight_accepts_total_or_amount_receipt_amounts(self):
        for field_name, value in (("total", "100.00"), ("amount", 100)):
            with self.subTest(field_name=field_name), TemporaryDirectory() as directory:
                root = Path(directory)
                state = self._valid_package_state(root)
                state["receipts"][0].pop("total_cents")
                state["receipts"][0][field_name] = value

                self.assertEqual(package_preflight(state, "quota-1"), [])

    def test_package_preflight_reports_invalid_receipt_amount_without_crashing(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["receipts"][0].pop("total_cents")
            state["receipts"][0]["total"] = "invalid"

            errors = package_preflight(state, "quota-1")

            self.assertIn("票据 receipt-1 金额格式无效", errors)

    def test_package_preflight_rejects_unknown_quota(self):
        with TemporaryDirectory() as directory:
            state = self._valid_package_state(Path(directory))

            errors = package_preflight(state, "missing-quota")

            self.assertIn("额度 missing-quota 不存在", errors)

    def test_package_preflight_rejects_quota_without_receipts(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["quotas"][0]["receipt_ids"] = []

            errors = package_preflight(state, "quota-1")

            self.assertIn("额度 quota-1 无分配票据", errors)

    def test_package_preflight_rejects_receipt_without_amount(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["receipts"][0].pop("total_cents")

            errors = package_preflight(state, "quota-1")

            self.assertIn("票据 receipt-1 缺少正金额", errors)

    def test_package_preflight_rejects_receipt_without_date(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["receipts"][0].pop("invoice_date")

            errors = package_preflight(state, "quota-1")

            self.assertIn("票据 receipt-1 缺少日期", errors)

    def test_package_preflight_rejects_receipt_with_missing_source_file(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["receipts"][0]["source_path"] = str(root / "missing.pdf")

            errors = package_preflight(state, "quota-1")

            self.assertIn("票据 receipt-1 源文件不存在", errors)

    def test_export_rejects_receipt_source_outside_workspace_imports(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory) / "workspace"
            workspace_dir.mkdir()
            state = self._valid_package_state(workspace_dir)
            outside_source = Path(directory) / "outside.pdf"
            outside_source.write_bytes(b"outside")
            state["receipts"][0]["source_path"] = str(outside_source)

            with self.assertRaisesRegex(ValueError, "imports"):
                export_quota_package(workspace_dir, state, "quota-1")

            self.assertFalse((workspace_dir / "exports").exists())

    def test_package_preflight_rejects_invalid_inventory_lines(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["lines_by_slot"]["quota-1"][0]["amount"] = "99.98"

            errors = package_preflight(state, "quota-1")

            self.assertTrue(any("物品明细校验失败" in error for error in errors))

    def test_package_preflight_rejects_invalid_template_directory(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["template_dir"] = str(root / "missing-templates")

            errors = package_preflight(state, "quota-1")

            self.assertTrue(any("模板无效" in error for error in errors))

    def test_package_preflight_accepts_iso_dates_and_cents_inventory_amounts(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            line = state["lines_by_slot"]["quota-1"][0]
            line.pop("unit_price")
            line.pop("amount")
            line["unit_price_cents"] = 5_000
            line["amount_cents"] = 10_000

            self.assertEqual(package_preflight(state, "quota-1"), [])

    def test_package_preflight_rejects_non_integer_quota_cents(self):
        for value in (10.5, True, float("nan"), float("inf")):
            with self.subTest(value=value), TemporaryDirectory() as directory:
                root = Path(directory)
                state = self._valid_package_state(root)
                state["quotas"][0]["capacity_cents"] = value

                errors = package_preflight(state, "quota-1")

                self.assertIn("额度 quota-1 上限必须是有限整数分", errors)

    def test_package_preflight_rejects_non_integer_receipt_cents(self):
        for value in (10.5, True, float("nan"), float("inf")):
            with self.subTest(value=value), TemporaryDirectory() as directory:
                root = Path(directory)
                state = self._valid_package_state(root)
                state["receipts"][0]["total_cents"] = value

                errors = package_preflight(state, "quota-1")

                self.assertIn("票据 receipt-1 total_cents 必须是有限整数分", errors)

    def test_package_preflight_rejects_non_integer_inventory_cents(self):
        for cents_key in ("unit_price_cents", "amount_cents"):
            for value in (10.5, True, float("nan"), float("inf")):
                with self.subTest(cents_key=cents_key, value=value), TemporaryDirectory() as directory:
                    root = Path(directory)
                    state = self._valid_package_state(root)
                    line = state["lines_by_slot"]["quota-1"][0]
                    line.pop("unit_price")
                    line.pop("amount")
                    line["unit_price_cents"] = 5_000
                    line["amount_cents"] = 10_000
                    line[cents_key] = value

                    errors = package_preflight(state, "quota-1")

                    self.assertTrue(any(f"{cents_key} 必须是有限整数分" in error for error in errors))

    def test_package_preflight_continues_after_invalid_cents_values(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            state = self._valid_package_state(root)
            state["quotas"][0]["capacity_cents"] = float("nan")
            state["receipts"][0]["total_cents"] = float("inf")
            state["receipts"][0]["source_path"] = str(root / "missing.pdf")

            errors = package_preflight(state, "quota-1")

            self.assertIn("额度 quota-1 上限必须是有限整数分", errors)
            self.assertIn("票据 receipt-1 total_cents 必须是有限整数分", errors)
            self.assertIn("票据 receipt-1 源文件不存在", errors)

    def test_export_quota_package_creates_a_clean_package_and_records_history(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory)
            state = self._valid_package_state(workspace_dir)

            result = export_quota_package(workspace_dir, state, "quota-1")

            package_dir = workspace_dir / "exports" / "quota-1-100.00"
            self.assertEqual(result, ExportResult(package_dir.resolve(), "quota-1", 10_000))
            self.assertEqual(
                {path.name for path in package_dir.iterdir()},
                {"入库单.xlsx", "出库单.xlsx", "原始票据"},
            )
            copied_receipt = package_dir / "原始票据" / "260713_01_100.00.pdf"
            self.assertEqual(copied_receipt.read_bytes(), b"receipt")
            self.assertEqual(
                state["exports"],
                [{
                    "path": "exports/quota-1-100.00",
                    "slot_id": "quota-1",
                    "total_cents": 10_000,
                    "created_at": state["exports"][0]["created_at"],
                }],
            )
            self.assertEqual(
                load_workspace(workspace_dir)["exports"],
                state["exports"],
            )

    def test_export_quota_package_uses_incremented_directory_without_overwriting(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory)
            state = self._valid_package_state(workspace_dir)

            first = export_quota_package(workspace_dir, state, "quota-1")
            first_inbound = (first.output_dir / "入库单.xlsx").read_bytes()
            second = export_quota_package(workspace_dir, state, "quota-1")

            self.assertEqual(first.output_dir.name, "quota-1-100.00")
            self.assertEqual(second.output_dir.name, "quota-1-100.00-02")
            self.assertEqual((first.output_dir / "入库单.xlsx").read_bytes(), first_inbound)
            self.assertEqual(len(state["exports"]), 2)

    def test_export_quota_package_preflight_failure_does_not_create_exports(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory)
            state = self._valid_package_state(workspace_dir)
            state["receipts"][0]["confirmed"] = False

            with self.assertRaisesRegex(ValueError, "票据 receipt-1 尚未人工确认"):
                export_quota_package(workspace_dir, state, "quota-1")

            self.assertFalse((workspace_dir / "exports").exists())
            self.assertEqual(state.get("exports"), None)

    def test_export_quota_package_copy_failure_cleans_staging_without_state_change(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory)
            state = self._valid_package_state(workspace_dir)

            with patch("core.materials_workspace.shutil.copy2", side_effect=OSError("copy failed")):
                with self.assertRaisesRegex(OSError, "copy failed"):
                    export_quota_package(workspace_dir, state, "quota-1")

            exports_dir = workspace_dir / "exports"
            self.assertEqual(list(exports_dir.iterdir()), [])
            self.assertNotIn("exports", state)

    def test_export_quota_package_save_failure_cleans_staging_without_state_change(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory)
            state = self._valid_package_state(workspace_dir)
            exports_dir = workspace_dir / "exports"

            def fail_save(_workspace_dir, _pending_state):
                self.assertTrue((exports_dir / "quota-1-100.00").is_dir())
                raise OSError("save failed")

            with patch("core.materials_workspace.save_workspace", side_effect=fail_save):
                with self.assertRaisesRegex(OSError, "save failed"):
                    export_quota_package(workspace_dir, state, "quota-1")

            self.assertEqual(list(exports_dir.iterdir()), [])
            self.assertNotIn("exports", state)

    def test_export_quota_package_rename_failure_leaves_no_history_or_package(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory)
            state = self._valid_package_state(workspace_dir)

            with patch("core.materials_workspace.Path.rename", side_effect=OSError("rename failed")):
                with self.assertRaisesRegex(OSError, "rename failed"):
                    export_quota_package(workspace_dir, state, "quota-1")

            exports_dir = workspace_dir / "exports"
            self.assertEqual(list(exports_dir.iterdir()), [])
            self.assertNotIn("exports", state)
            self.assertFalse((workspace_dir / "workspace.json").exists())

    def test_create_save_and_reload_workspace(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workspace_dir, state = create_workspace(root, "Project Alpha", "templates")
            state["step"] = "review"
            state["lines_by_slot"] = {"quota-1": [{"name": "sensor", "amount_cents": 1200}]}
            state["exports"] = [{"path": "exports/package-1", "created_at": "2026-07-13T10:00:00"}]

            save_workspace(workspace_dir, state)
            loaded = load_workspace(workspace_dir)

            self.assertEqual(workspace_dir.parent, (root / "workspaces").resolve())
            self.assertTrue((workspace_dir / "workspace.json").is_file())
            for directory_name in ("imports", "records", "exports", "backups"):
                self.assertTrue((workspace_dir / directory_name).is_dir())
            self.assertEqual(loaded["name"], "Project Alpha")
            self.assertEqual(loaded["template_dir"], "templates")
            self.assertEqual(loaded["step"], "review")
            self.assertEqual(loaded["lines_by_slot"], state["lines_by_slot"])
            self.assertEqual(loaded["exports"], state["exports"])
            self.assertRegex(workspace_dir.name, r"^[A-Za-z0-9_-]+-[0-9a-f]{32}$")

    def test_receipt_and_quota_dataclasses_round_trip(self):
        with TemporaryDirectory() as directory:
            workspace_dir, state = create_workspace(directory, "Roundtrip")
            state["receipts"] = [
                Receipt(
                    "receipt-1", "receipts/original.pdf", date(2026, 7, 13), 12345,
                    vendor_name="Vendor", invoice_number="INV-1", ocr_text="text",
                    is_material=False, extraction_note="checked",
                )
            ]
            state["quotas"] = [QuotaSlot("quota-1", 50000, "Materials")]

            save_workspace(workspace_dir, state)
            loaded = load_workspace(workspace_dir)

            self.assertEqual(workspace_receipts(loaded), state["receipts"])
            self.assertEqual(workspace_quotas(loaded), state["quotas"])
            with (workspace_dir / "workspace.json").open(encoding="utf-8") as handle:
                stored = json.load(handle)
            self.assertEqual(stored["receipts"][0]["invoice_date"], "2026-07-13")
            self.assertEqual(stored["receipts"][0]["total_cents"], 12345)

    def test_extended_receipt_and_quota_fields_round_trip(self):
        with TemporaryDirectory() as directory:
            workspace_dir, state = create_workspace(directory, "Extended fields")
            state["receipts"] = [{
                "receipt_id": "receipt-1",
                "source_path": "imports/original.pdf",
                "invoice_date": "2026-07-13",
                "total_cents": 12345,
                "confirmed": "false",
                "wizard_note": "needs review",
            }]
            state["quotas"] = [{
                "slot_id": "quota-1",
                "capacity_cents": 50000,
                "receipt_ids": ["receipt-1"],
                "wizard_step": "allocated",
            }]

            save_workspace(workspace_dir, state)
            loaded = load_workspace(workspace_dir)

            self.assertFalse(loaded["receipts"][0]["confirmed"])
            self.assertEqual(loaded["receipts"][0]["wizard_note"], "needs review")
            self.assertEqual(loaded["quotas"][0]["receipt_ids"], ["receipt-1"])
            self.assertEqual(loaded["quotas"][0]["wizard_step"], "allocated")
            self.assertEqual(workspace_receipts(loaded)[0].receipt_id, "receipt-1")
            self.assertEqual(workspace_quotas(loaded)[0].slot_id, "quota-1")

    def test_datetime_receipt_date_is_saved_as_date_only(self):
        with TemporaryDirectory() as directory:
            workspace_dir, state = create_workspace(directory, "Datetime")
            state["receipts"] = [Receipt(
                "receipt-1", "imports/original.pdf", datetime(2026, 7, 13, 14, 30), 12345
            )]

            save_workspace(workspace_dir, state)

            loaded = load_workspace(workspace_dir)

            self.assertEqual(loaded["receipts"][0]["invoice_date"], "2026-07-13")
            self.assertEqual(workspace_receipts(loaded)[0].invoice_date, date(2026, 7, 13))

    def test_invalid_workspace_name_is_safe_and_falls_back_when_empty(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workspace_dir, state = create_workspace(root, "../unsafe name!?")
            fallback_dir, fallback_state = create_workspace(root, "")

            self.assertEqual(workspace_dir.parent, (root / "workspaces").resolve())
            self.assertNotIn("..", workspace_dir.name)
            self.assertRegex(workspace_dir.name, r"^[A-Za-z0-9_-]+$")
            self.assertTrue(fallback_dir.name.startswith("materials-"))
            self.assertEqual(state["name"], "../unsafe name!?")
            self.assertEqual(fallback_state["name"], "")

    def test_migrate_legacy_workspace_copies_all_workspace_data_once(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            legacy_files = {
                "imports/nested/receipt.pdf": b"receipt",
                "records/audit/entry.json": b'{"event":"import"}',
                "exports/historical-package/manifest.txt": b"export history",
                "backups/2026-07-13/workspace.json": b"backup",
            }
            for relative_path, contents in legacy_files.items():
                path = root / relative_path
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(contents)
            (root / "workspace.json").write_text(
                json.dumps({
                    "name": "Legacy",
                    "receipts": [{
                        "receipt_id": "legacy-receipt",
                        "source_path": "imports/nested/receipt.pdf",
                    }],
                    "quotas": [],
                    "exports": [{"path": "exports/historical-package"}],
                }),
                encoding="utf-8",
            )

            migrated = migrate_legacy_workspace(root)
            repeated = migrate_legacy_workspace(root)

            self.assertIsNotNone(migrated)
            self.assertEqual(repeated, migrated)
            self.assertEqual(len(list((root / "workspaces").iterdir())), 1)
            for relative_path, contents in legacy_files.items():
                self.assertEqual((migrated / relative_path).read_bytes(), contents)
            state = load_workspace(migrated)
            self.assertEqual(state["exports"][0]["path"], "exports/historical-package")
            self.assertTrue((migrated / state["exports"][0]["path"]).is_dir())
            self.assertEqual(
                Path(state["receipts"][0]["source_path"]),
                migrated / "imports/nested/receipt.pdf",
            )

    def test_migrate_legacy_workspace_copy_failure_preserves_legacy_root_and_cleans_staging(self):
        self._assert_legacy_migration_failure("core.materials_workspace.shutil.copytree")

    def test_migrate_legacy_workspace_save_failure_preserves_legacy_root_and_cleans_staging(self):
        self._assert_legacy_migration_failure("core.materials_workspace.save_workspace")

    def test_migrate_legacy_workspace_rename_failure_preserves_legacy_root_and_cleans_staging(self):
        self._assert_legacy_migration_failure("core.materials_workspace.Path.rename")

    def test_bad_json_raises_readable_value_error(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory) / "workspace"
            workspace_dir.mkdir()
            (workspace_dir / "workspace.json").write_text("{not valid", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "workspace"):
                load_workspace(workspace_dir)

    def test_missing_workspace_file_raises_readable_value_error(self):
        with TemporaryDirectory() as directory:
            workspace_dir = Path(directory) / "workspace"
            workspace_dir.mkdir()

            with self.assertRaisesRegex(ValueError, "Workspace file is missing"):
                load_workspace(workspace_dir)

    def test_load_workspace_rejects_bad_schema_or_collection_types_without_rewriting_file(self):
        invalid_states = (
            {"schema_version": 99, "receipts": [], "quotas": [], "lines_by_slot": {}, "exports": []},
            {"schema_version": 1, "receipts": {}, "quotas": [], "lines_by_slot": {}, "exports": []},
            {"schema_version": 1, "receipts": [], "quotas": {}, "lines_by_slot": {}, "exports": []},
            {"schema_version": 1, "receipts": [], "quotas": [], "lines_by_slot": [], "exports": []},
            {"schema_version": 1, "receipts": [], "quotas": [], "lines_by_slot": {}, "exports": {}},
        )
        for state in invalid_states:
            with self.subTest(state=state), TemporaryDirectory() as directory:
                workspace_dir = Path(directory) / "workspace"
                workspace_dir.mkdir()
                path = workspace_dir / "workspace.json"
                path.write_text(json.dumps(state), encoding="utf-8")
                original = path.read_bytes()

                with self.assertRaises(ValueError):
                    load_workspace(workspace_dir)

                self.assertEqual(path.read_bytes(), original)

    def test_false_boolean_values_remain_false(self):
        with TemporaryDirectory() as directory:
            workspace_dir, state = create_workspace(directory, "Boolean values")
            state["receipts"] = [
                {"receipt_id": "native", "confirmed": False},
                {"receipt_id": "string", "confirmed": "false"},
            ]

            save_workspace(workspace_dir, state)
            loaded = load_workspace(workspace_dir)

            self.assertEqual([receipt["confirmed"] for receipt in loaded["receipts"]], [False, False])

    def test_repeated_saves_preserve_state_and_original_receipts(self):
        with TemporaryDirectory() as directory:
            workspace_dir, state = create_workspace(directory, "Repeat")
            receipt_file = workspace_dir / "receipts" / "original.pdf"
            receipt_file.parent.mkdir()
            receipt_file.write_bytes(b"original receipt")
            state["receipts"] = [Receipt("receipt-1", str(receipt_file), None, None)]

            save_workspace(workspace_dir, state)
            state["step"] = "export"
            save_workspace(workspace_dir, state)

            self.assertEqual(load_workspace(workspace_dir)["step"], "export")
            self.assertEqual(receipt_file.read_bytes(), b"original receipt")

    def test_export_and_restore_workspace_backup_preserves_state_and_imports(self):
        with TemporaryDirectory() as directory, TemporaryDirectory() as restore_directory:
            root = Path(directory)
            workspace_dir, state = create_workspace(root, "Project Alpha")
            receipt = workspace_dir / "imports" / "receipt.pdf"
            receipt.write_bytes(b"receipt")
            state["receipts"] = [{"receipt_id": "receipt-1", "source_path": str(receipt)}]
            save_workspace(workspace_dir, state)
            (workspace_dir / "backups" / "old-backup.zip").write_bytes(b"old backup")

            archive = export_workspace_backup(workspace_dir)
            restored = restore_workspace_backup(restore_directory, archive)

            self.assertEqual(load_workspace(restored)["name"], "Project Alpha")
            self.assertEqual((restored / "imports" / "receipt.pdf").read_bytes(), b"receipt")
            self.assertEqual(restored.parent, (Path(restore_directory) / "workspaces").resolve())
            with ZipFile(archive) as backup:
                self.assertNotIn("backups/old-backup.zip", backup.namelist())
                self.assertTrue(
                    all(
                        member.compress_type == ZIP_DEFLATED
                        for member in backup.infolist()
                        if not member.is_dir()
                    )
                )

    def test_export_workspace_backup_rejects_snapshots_restore_would_reject(self):
        with TemporaryDirectory() as directory:
            workspace_dir, _ = create_workspace(directory, "Oversized")

            with patch("core.materials_workspace.MAX_RESTORE_TOTAL_BYTES", 1):
                with self.assertRaisesRegex(ValueError, "total size"):
                    export_workspace_backup(workspace_dir)

            self.assertFalse(list((workspace_dir / "backups").glob("*.zip")))

    def test_restore_workspace_backup_rejects_traversal_and_cleans_staging(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "traversal.zip"
            with ZipFile(archive, "w") as backup:
                backup.writestr("workspace.json", "{}")
                backup.writestr("../escape.txt", "unsafe")

            with self.assertRaisesRegex(ValueError, "unsafe path"):
                restore_workspace_backup(root, archive)

            workspaces = root / "workspaces"
            self.assertFalse((root / "escape.txt").exists())
            self.assertFalse(workspaces.exists() and list(workspaces.iterdir()))

    def test_restore_workspace_backup_requires_root_workspace_file(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "missing-workspace.zip"
            with ZipFile(archive, "w") as backup:
                backup.writestr("imports/receipt.pdf", b"receipt")

            with self.assertRaisesRegex(ValueError, "workspace.json"):
                restore_workspace_backup(root, archive)

    def test_restore_workspace_backup_rejects_absolute_and_drive_members(self):
        for member_name in ("/absolute.txt", "C:/drive.txt"):
            with self.subTest(member_name=member_name), TemporaryDirectory() as directory:
                root = Path(directory)
                archive = root / "unsafe.zip"
                with ZipFile(archive, "w") as backup:
                    backup.writestr("workspace.json", "{}")
                    backup.writestr(member_name, "unsafe")

                with self.assertRaisesRegex(ValueError, "unsafe path"):
                    restore_workspace_backup(root, archive)

    def test_restore_workspace_backup_rewrites_receipt_paths_for_export(self):
        with TemporaryDirectory() as directory, TemporaryDirectory() as restore_directory:
            root = Path(directory)
            workspace_dir, state = create_workspace(root, "Project Alpha")
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单模板.xlsx")
            self._write_template(templates / "出库单模板.xlsx")
            receipt = workspace_dir / "imports" / "nested" / "receipt.pdf"
            receipt.parent.mkdir()
            receipt.write_bytes(b"receipt")
            state.update({
                "template_dir": str(templates),
                "receipts": [{
                    "receipt_id": "receipt-1", "source_path": str(receipt),
                    "invoice_date": "2026-07-13", "total_cents": 10_000,
                    "is_material": True, "confirmed": True,
                }],
                "quotas": [{
                    "slot_id": "quota-1", "capacity_cents": 10_000,
                    "receipt_ids": ["receipt-1"],
                }],
                "lines_by_slot": {"quota-1": [{
                    "inventory_date": "2026-07-13", "name": "Sensor",
                    "specification": "S-1", "unit": "piece", "quantity": "2",
                    "unit_price": "50.00", "amount": "100.00", "supplier_name": "Vendor",
                }]},
            })
            save_workspace(workspace_dir, state)
            archive = export_workspace_backup(workspace_dir)
            receipt.unlink()

            restored = restore_workspace_backup(restore_directory, archive)
            restored_state = load_workspace(restored)

            self.assertEqual(
                Path(restored_state["receipts"][0]["source_path"]),
                restored / "imports" / "nested" / "receipt.pdf",
            )
            result = export_quota_package(restored, restored_state, "quota-1")
            self.assertEqual(
                (result.output_dir / "原始票据" / "260713_01_100.00.pdf").read_bytes(),
                b"receipt",
            )

    def test_restore_workspace_backup_rejects_receipt_without_restored_import(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workspace_dir, state = create_workspace(root, "Project Alpha")
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": str(root / "outside.pdf"),
            }]
            save_workspace(workspace_dir, state)
            archive = export_workspace_backup(workspace_dir)

            with self.assertRaisesRegex(ValueError, "receipt source"):
                restore_workspace_backup(root / "restored-home", archive)

            workspaces = root / "restored-home" / "workspaces"
            self.assertFalse(workspaces.exists() and list(workspaces.iterdir()))

    def test_restore_workspace_backup_cleans_staging_for_truncated_zip(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "truncated.zip"
            archive.write_bytes(b"PK\x03\x04truncated")

            with self.assertRaises(ValueError):
                restore_workspace_backup(root, archive)

            workspaces = root / "workspaces"
            self.assertFalse(workspaces.exists() and list(workspaces.glob(".restore-*.staging")))

    def test_restore_workspace_backup_cleans_staging_when_archive_read_fails(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "valid.zip"
            with ZipFile(archive, "w") as backup:
                backup.writestr("workspace.json", "{}")

            with patch("core.materials_workspace.ZipFile.open", side_effect=BadZipFile("broken")):
                with self.assertRaisesRegex(ValueError, "valid ZIP"):
                    restore_workspace_backup(root, archive)

            workspaces = root / "workspaces"
            self.assertFalse(workspaces.exists() and list(workspaces.glob(".restore-*.staging")))

    def test_restore_workspace_backup_rejects_archives_with_too_many_entries(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "many-entries.zip"
            with ZipFile(archive, "w") as backup:
                backup.writestr("workspace.json", "{}")
                backup.writestr("imports/receipt.pdf", b"receipt")

            with patch("core.materials_workspace.MAX_RESTORE_ENTRIES", 1, create=True):
                with self.assertRaisesRegex(ValueError, "too many entries"):
                    restore_workspace_backup(root, archive)

            workspaces = root / "workspaces"
            self.assertFalse(workspaces.exists() and list(workspaces.iterdir()))

    def test_restore_workspace_backup_rejects_declared_member_size_above_limit(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "large-member.zip"
            with ZipFile(archive, "w") as backup:
                backup.writestr("workspace.json", "{}")
                backup.writestr("imports/receipt.pdf", b"too large")

            with patch("core.materials_workspace.MAX_RESTORE_MEMBER_BYTES", 4, create=True):
                with self.assertRaisesRegex(ValueError, "member too large"):
                    restore_workspace_backup(root, archive)

            workspaces = root / "workspaces"
            self.assertFalse(workspaces.exists() and list(workspaces.iterdir()))

    def test_restore_workspace_backup_rejects_actual_stream_size_above_total_limit(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            archive = root / "streamed-large.zip"
            with ZipFile(archive, "w") as backup:
                backup.writestr("workspace.json", "{}")
                backup.writestr("imports/receipt.pdf", b"x")

            original_open = ZipFile.open

            def open_with_oversized_payload(backup, member, *args, **kwargs):
                if getattr(member, "filename", member) == "imports/receipt.pdf":
                    return BytesIO(b"x" * 16)
                return original_open(backup, member, *args, **kwargs)

            with patch("core.materials_workspace.MAX_RESTORE_TOTAL_BYTES", 8, create=True), patch(
                "core.materials_workspace.ZipFile.open", autospec=True,
                side_effect=open_with_oversized_payload,
            ):
                with self.assertRaisesRegex(ValueError, "total size"):
                    restore_workspace_backup(root, archive)

            workspaces = root / "workspaces"
            self.assertFalse(workspaces.exists() and list(workspaces.iterdir()))

    @staticmethod
    def _valid_package_state(root: Path) -> dict:
        templates = root / "templates"
        templates.mkdir()
        MaterialsWorkspaceTests._write_template(templates / "入库单_模板.xlsx")
        MaterialsWorkspaceTests._write_template(templates / "出库单_模板.xlsx")
        source = root / "imports" / "receipt.pdf"
        source.parent.mkdir(exist_ok=True)
        source.write_bytes(b"receipt")
        return {
            "template_dir": str(templates),
            "receipts": [{
                "receipt_id": "receipt-1",
                "source_path": str(source),
                "invoice_date": "2026-07-13",
                "total_cents": 10_000,
                "is_material": True,
                "confirmed": True,
            }],
            "quotas": [{
                "slot_id": "quota-1",
                "capacity_cents": 10_000,
                "receipt_ids": ["receipt-1"],
            }],
            "lines_by_slot": {"quota-1": [{
                "inventory_date": "2026-07-13",
                "name": "开发板",
                "specification": "F4",
                "unit": "块",
                "quantity": "2",
                "unit_price": "50.00",
                "amount": "100.00",
            }]},
        }

    @staticmethod
    def _write_template(
        path: Path,
        headers=None,
        total_row: int = 7,
        total_label_column: int = 6,
        amount_formula: bool = True,
        capitalized_total_formula: bool = True,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        if headers is None:
            headers = [
                "入库日期" if "入库单" in path.stem else "出库日期",
                "产品名称",
                "规格型号",
                "单位",
                "入库数量" if "入库单" in path.stem else "出库数量",
                "单价(元)",
                "金额(元)",
                "供货单位" if "入库单" in path.stem else "用途",
                "经办人" if "入库单" in path.stem else "领用人",
                "管理员",
            ]
        for column, header in enumerate(headers, start=1):
            sheet.cell(4, column).value = header
        sheet.cell(total_row, total_label_column).value = "合计："
        sheet.cell(total_row, 7).value = f"=SUM(G5:G{total_row - 1})" if amount_formula else None
        sheet.cell(total_row, 8).value = "大写："
        sheet.cell(total_row, 9).value = (
            f'=TEXT(G{total_row},"[DBNUM2]")&"元整"' if capitalized_total_formula else None
        )
        try:
            workbook.save(path)
        finally:
            workbook.close()

    def _assert_legacy_migration_failure(self, patch_target: str) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            receipt_path = root / "imports" / "receipt.pdf"
            receipt_path.parent.mkdir()
            receipt_path.write_bytes(b"legacy receipt")
            state_path = root / "workspace.json"
            state_path.write_text(json.dumps({"name": "Legacy", "receipts": [], "quotas": []}), encoding="utf-8")
            original_state = state_path.read_bytes()
            original_receipt = receipt_path.read_bytes()

            with patch(patch_target, side_effect=OSError("migration failed")):
                with self.assertRaisesRegex(OSError, "migration failed"):
                    migrate_legacy_workspace(root)

            workspaces_dir = root / "workspaces"
            self.assertEqual(state_path.read_bytes(), original_state)
            self.assertEqual(receipt_path.read_bytes(), original_receipt)
            self.assertFalse(any(path.name.startswith("legacy-") for path in workspaces_dir.iterdir()))
            self.assertEqual(list(workspaces_dir.glob(".legacy-*.staging")), [])


if __name__ == "__main__":
    unittest.main()
