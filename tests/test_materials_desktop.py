from contextlib import contextmanager, redirect_stdout
from io import BytesIO, StringIO
import gc
import os
from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import MagicMock, patch
import warnings
from zipfile import ZipFile

from openpyxl import Workbook
import run_materials_desktop

try:
    from materials_desktop import create_materials_app
    from core.materials_assistant import Receipt, ingest_receipt
    from core.inventory_line_extraction import RecognizedLineItem
    from core.ocr import OCRResult, get_ocr_provider
    from core.materials_workspace import load_workspace, save_workspace
except ModuleNotFoundError:
    create_materials_app = None


@unittest.skipUnless(create_materials_app, "Flask desktop dependencies are not installed")
class MaterialsDesktopTests(unittest.TestCase):
    _REQUIRED_OCR_MODEL_PATHS = (
        Path("whl/det/ch/ch_PP-OCRv4_det_infer"),
        Path("whl/rec/ch/ch_PP-OCRv4_rec_infer"),
        Path("whl/cls/ch_ppocr_mobile_v2.0_cls_infer"),
    )

    @contextmanager
    def _draft_refresh_workspace(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source1 = workspace_dir / "imports" / "r1.pdf"
            source2 = workspace_dir / "imports" / "r2.pdf"
            source1.write_bytes(b"receipt-1")
            source2.write_bytes(b"receipt-2")
            state = load_workspace(workspace_dir)
            state["receipts"] = [
                {
                    "receipt_id": "r1", "source_path": str(source1),
                    "invoice_date": "2026-07-15", "total_cents": 10_000,
                    "vendor_name": "供应商", "confirmed": True, "is_material": True,
                },
                {
                    "receipt_id": "r2", "source_path": str(source2),
                    "invoice_date": "2026-07-15", "total_cents": 8_000,
                    "vendor_name": "供应商", "confirmed": True, "is_material": True,
                },
            ]
            state["quotas"] = [{
                "slot_id": "q1", "capacity_cents": 20_000, "receipt_ids": ["r1"],
            }]
            state["lines_by_slot"] = {"q1": [{
                "inventory_date": "2026-07-15", "name": "旧名称", "specification": "",
                "unit": "件", "quantity": "1", "unit_price": "100.00",
                "amount": "100.00", "supplier_name": "供应商",
                "recognition_status": "已校验", "_auto_generated": True,
                "_source_receipt_id": "r1",
            }]}
            save_workspace(workspace_dir, state)
            yield client, workspace_id, workspace_dir

    @classmethod
    def _write_required_ocr_models(cls, models: Path) -> None:
        for relative_path in cls._REQUIRED_OCR_MODEL_PATHS:
            model_path = models / relative_path
            model_path.mkdir(parents=True)
            (model_path / "inference.pdmodel").write_bytes(b"model")
            (model_path / "inference.pdiparams").write_bytes(b"parameters")

    def test_headless_smoke_configures_assets_and_returns_exact_marker(self):
        self.assertTrue(hasattr(run_materials_desktop, "run_smoke_test"))
        self.assertTrue(hasattr(run_materials_desktop, "LOCAL_MATERIALS_SMOKE_OK"))
        self.assertTrue(hasattr(run_materials_desktop, "_initialize_paddle_ocr"))
        with TemporaryDirectory() as directory:
            app = create_materials_app(directory)
            models = Path(directory) / "models"
            with patch(
                "run_materials_desktop.configure_release_assets", return_value=models
            ) as configure, patch(
                "run_materials_desktop._initialize_paddle_ocr"
            ) as initialize, patch("run_materials_desktop.create_materials_app", return_value=app):
                marker = run_materials_desktop.run_smoke_test()

        configure.assert_called_once_with()
        initialize.assert_called_once_with(models)
        self.assertEqual(marker, "LOCAL_MATERIALS_SMOKE_OK")

    def test_smoke_main_does_not_start_desktop_window(self):
        self.assertTrue(hasattr(run_materials_desktop, "_start_desktop"))
        with patch(
            "run_materials_desktop.run_smoke_test",
            return_value="LOCAL_MATERIALS_SMOKE_OK",
        ), patch("run_materials_desktop._start_desktop") as start_desktop:
            stdout = StringIO()
            with redirect_stdout(stdout):
                status = run_materials_desktop.main(["--smoke"])

        self.assertEqual(status, 0)
        self.assertEqual(stdout.getvalue(), "LOCAL_MATERIALS_SMOKE_OK\n")
        start_desktop.assert_not_called()

    def test_ocr_worker_main_returns_before_starting_desktop(self):
        with patch("run_materials_desktop.run_worker", return_value=0) as run_worker, patch(
            "run_materials_desktop._start_desktop"
        ) as start_desktop, patch(
            "run_materials_desktop.create_materials_app"
        ) as create_app:
            status = run_materials_desktop.main(
                ["--ocr-worker", "scan.png", "result.json"]
            )

        self.assertEqual(status, 0)
        run_worker.assert_called_once_with("scan.png", "result.json")
        start_desktop.assert_not_called()
        create_app.assert_not_called()

    def test_desktop_window_uses_local_http_server_instead_of_wsgi_object(self):
        class ClosedEvent:
            def __init__(self):
                self.handlers = []

            def __iadd__(self, handler):
                self.handlers.append(handler)
                return self

        class Events:
            def __init__(self):
                self.closed = ClosedEvent()

        app = object()
        webview = MagicMock()
        window = MagicMock()
        window.events = Events()
        webview.create_window.return_value = window
        server = MagicMock()
        server.server_port = 51337

        with patch.dict(sys.modules, {"webview": webview}), patch(
            "run_materials_desktop.make_server", return_value=server, create=True
        ) as make_server, patch("run_materials_desktop.Thread", create=True) as thread:
            run_materials_desktop._start_desktop(app)

        make_server.assert_called_once_with("127.0.0.1", 0, app)
        thread.assert_called_once_with(target=server.serve_forever, daemon=False)
        webview.create_window.assert_called_once_with(
            "大创报销材料助手",
            "http://127.0.0.1:51337/",
            width=1280,
            height=860,
            min_size=(980, 680),
        )
        webview.start.assert_called_once_with()
        server.shutdown.assert_not_called()
        self.assertEqual(window.events.closed.handlers, [server.shutdown])

    def test_storage_error_on_homepage_is_not_redirected_back_to_homepage(self):
        with TemporaryDirectory() as directory:
            client = create_materials_app(directory).test_client()
            with patch("materials_desktop.render_template", side_effect=OSError("disk unavailable")):
                response = client.get("/")

        self.assertEqual(response.status_code, 500)
        self.assertIn("本地文件写入失败", response.get_data(as_text=True))

    def test_app_uses_an_absolute_template_directory_from_bundle_root(self):
        project_root = Path(__file__).resolve().parents[1]
        with TemporaryDirectory() as directory, patch(
            "materials_desktop.bundle_root", return_value=project_root
        ), patch(
            "materials_desktop.bundled_ocr_models_path", return_value=Path(directory) / "missing-models"
        ):
            app = create_materials_app(directory)

        self.assertEqual(Path(app.template_folder), project_root / "app" / "templates")

    def test_new_workspace_redirects_to_first_wizard_step(self):
        with TemporaryDirectory() as directory:
            client = create_materials_app(directory).test_client()

            response = client.post("/workspace/new", data={"name": "Project Alpha"})

            self.assertEqual(response.status_code, 302)
            self.assertRegex(response.headers["Location"], r"/workspace/[^/]+/step/1$")

    def test_workspace_can_be_deleted_from_the_workspace_list(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)

            response = client.post(f"/workspace/{workspace_id}/delete", follow_redirects=True)

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.request.path, "/")
            self.assertFalse((root / "workspaces" / workspace_id).exists())
            self.assertIn("已删除工作区", response.get_data(as_text=True))

    def test_workspace_can_be_renamed_without_changing_its_internal_id(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id

            response = client.post(
                f"/workspace/{workspace_id}/rename",
                data={"name": "2027 大创材料"},
                follow_redirects=True,
            )

            self.assertEqual(response.status_code, 200)
            self.assertTrue(workspace_dir.is_dir())
            self.assertEqual(load_workspace(workspace_dir)["name"], "2027 大创材料")
            self.assertIn("2027 大创材料", response.get_data(as_text=True))

    def test_blank_workspace_name_is_rejected_without_changing_the_name(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id

            response = client.post(
                f"/workspace/{workspace_id}/rename",
                data={"name": "   "},
                follow_redirects=True,
            )

            self.assertEqual(load_workspace(workspace_dir)["name"], "Project Alpha")
            self.assertIn("工作区名称不能为空", response.get_data(as_text=True))

    def test_index_shows_workspace_business_name_instead_of_internal_id(self):
        with TemporaryDirectory() as directory:
            client = create_materials_app(directory).test_client()
            workspace_id = self._new_workspace_id(client)

            page = client.get("/").get_data(as_text=True)

            self.assertIn("Project Alpha", page)
            self.assertNotIn(f">{workspace_id}<", page)

    def test_corrupted_workspace_is_listed_without_redirecting_index(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            broken = root / "workspaces" / "broken-id"
            broken.mkdir(parents=True)
            (broken / "workspace.json").write_text("{broken", encoding="utf-8")
            client = create_materials_app(root).test_client()

            response = client.get("/", follow_redirects=False)

            self.assertEqual(response.status_code, 200)
            self.assertIn("无法读取工作区", response.get_data(as_text=True))

    def test_corrupted_workspace_can_still_be_deleted(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            broken = root / "workspaces" / "broken-id"
            broken.mkdir(parents=True)
            (broken / "workspace.json").write_text("{broken", encoding="utf-8")
            client = create_materials_app(root).test_client()

            response = client.post(
                "/workspace/broken-id/delete",
                follow_redirects=True,
            )

            self.assertEqual(response.status_code, 200)
            self.assertFalse(broken.exists())
            self.assertIn("已删除工作区", response.get_data(as_text=True))

    def test_new_workspace_uses_bundled_templates_and_can_open_import_step(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            templates = root / "inventory_templates"
            templates.mkdir()
            self._write_template(templates / "入库单.xlsx")
            self._write_template(templates / "出库单.xlsx")
            with patch(
                "materials_desktop.bundled_inventory_templates_path", return_value=templates
            ), patch(
                "materials_desktop.bundled_ocr_models_path", return_value=root / "missing-models"
            ):
                client = create_materials_app(root).test_client()
                workspace_id = self._new_workspace_id(client)

            state = load_workspace(root / "workspaces" / workspace_id)
            self.assertEqual(state["template_dir"], str(templates.resolve()))
            self.assertEqual(client.get(f"/workspace/{workspace_id}/step/2").status_code, 200)

    def test_custom_template_upload_copies_a_valid_pair_into_the_workspace(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            custom = root / "custom"
            custom.mkdir()
            inbound = custom / "学校入库单.xlsx"
            outbound = custom / "学校出库单.xlsx"
            self._write_template(inbound)
            self._write_template(outbound)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id

            response = client.post(
                f"/workspace/{workspace_id}/templates",
                data={
                    "template_mode": "custom",
                    "inbound_template": (BytesIO(inbound.read_bytes()), inbound.name),
                    "outbound_template": (BytesIO(outbound.read_bytes()), outbound.name),
                },
                content_type="multipart/form-data",
            )
            saved = load_workspace(workspace_dir)

            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/step/2"))
            self.assertEqual(Path(saved["template_dir"]), workspace_dir / "templates")
            self.assertTrue((workspace_dir / "templates" / "入库单.xlsx").is_file())
            self.assertTrue((workspace_dir / "templates" / "出库单.xlsx").is_file())

    def test_invalid_custom_template_upload_keeps_the_bundled_templates(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            original_template_dir = load_workspace(workspace_dir)["template_dir"]

            response = client.post(
                f"/workspace/{workspace_id}/templates",
                data={
                    "template_mode": "custom",
                    "inbound_template": (BytesIO(b"not an xlsx"), "入库表.xlsx"),
                    "outbound_template": (BytesIO(b"not an xlsx"), "出库表.xlsx"),
                },
                content_type="multipart/form-data",
                follow_redirects=True,
            )

            self.assertEqual(load_workspace(workspace_dir)["template_dir"], original_template_dir)
            self.assertIn("自定义模板未启用", response.get_data(as_text=True))
            self.assertFalse((workspace_dir / "templates").exists())

    def test_workspace_can_switch_from_custom_templates_back_to_bundled_templates(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            custom = root / "custom"
            custom.mkdir()
            inbound = custom / "入库单.xlsx"
            outbound = custom / "出库单.xlsx"
            self._write_template(inbound)
            self._write_template(outbound)
            app = create_materials_app(root)
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            client.post(
                f"/workspace/{workspace_id}/templates",
                data={
                    "template_mode": "custom",
                    "inbound_template": (BytesIO(inbound.read_bytes()), inbound.name),
                    "outbound_template": (BytesIO(outbound.read_bytes()), outbound.name),
                },
                content_type="multipart/form-data",
            )

            response = client.post(
                f"/workspace/{workspace_id}/templates",
                data={"template_mode": "bundled"},
            )

            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/step/1"))
            self.assertEqual(
                load_workspace(workspace_dir)["template_dir"],
                str(app.config["MATERIALS_TEMPLATE_DIR"]),
            )

    def test_app_configures_bundled_ocr_assets_only_when_models_exist(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            bundle = root / "bundle"
            models = bundle / "ocr_models"
            with patch("materials_desktop.bundled_ocr_models_path", return_value=models), patch(
                "materials_desktop.bundle_root", return_value=bundle
            ), patch("materials_desktop.configure_bundled_ocr_assets") as configure:
                create_materials_app(root / "without-models")
                configure.assert_not_called()

                self._write_required_ocr_models(models)
                home = root / "with-models"
                create_materials_app(home)

            configure.assert_called_once_with(bundle, home)

    def test_app_without_bundled_models_forces_mock_ocr_and_never_calls_paddle(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            with patch(
                "materials_desktop.bundled_ocr_models_path", return_value=root / "missing-models"
            ), patch("materials_desktop.ingest_receipt", wraps=ingest_receipt) as ingest, patch(
                "core.ocr.get_ocr_provider", wraps=get_ocr_provider
            ) as get_provider, patch("core.ocr.paddle_provider.PaddleOCRProvider") as paddle:
                app = create_materials_app(root)
                client = app.test_client()
                workspace_id = self._new_workspace_id(client)
                self._set_valid_template(root, root / "workspaces" / workspace_id)

                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"ocr_provider": "paddle", "receipts": (BytesIO(b"receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                    follow_redirects=True,
                )

            self.assertEqual(app.config["MATERIALS_OCR_PROVIDER"], "mock")
            self.assertTrue(app.config["MATERIALS_OFFLINE_OCR"])
            self.assertEqual(response.status_code, 200)
            self.assertIn("手工确认", response.get_data(as_text=True))
            self.assertEqual(ingest.call_args.args[1], "mock")
            get_provider.assert_called_once_with("mock")
            paddle.assert_not_called()

    def test_app_with_partial_bundled_models_forces_mock_ocr_and_never_calls_paddle(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            models = root / "partial-models"
            (models / self._REQUIRED_OCR_MODEL_PATHS[0]).mkdir(parents=True)
            with patch("materials_desktop.bundled_ocr_models_path", return_value=models), patch(
                "materials_desktop.ingest_receipt", wraps=ingest_receipt
            ) as ingest, patch("core.ocr.get_ocr_provider", wraps=get_ocr_provider) as get_provider, patch(
                "core.ocr.paddle_provider.PaddleOCRProvider"
            ) as paddle:
                app = create_materials_app(root)
                client = app.test_client()
                workspace_id = self._new_workspace_id(client)
                self._set_valid_template(root, root / "workspaces" / workspace_id)
                client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"ocr_provider": "paddle", "receipts": (BytesIO(b"receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                )

            self.assertEqual(app.config["MATERIALS_OCR_PROVIDER"], "mock")
            self.assertTrue(app.config["MATERIALS_OFFLINE_OCR"])
            self.assertEqual(ingest.call_args.args[1], "mock")
            get_provider.assert_called_once_with("mock")
            paddle.assert_not_called()

    def test_app_with_empty_model_directories_forces_mock_ocr_and_never_calls_paddle(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            models = root / "empty-models"
            for relative_path in self._REQUIRED_OCR_MODEL_PATHS:
                (models / relative_path).mkdir(parents=True)
            with patch("materials_desktop.bundled_ocr_models_path", return_value=models), patch(
                "materials_desktop.ingest_receipt", wraps=ingest_receipt
            ) as ingest, patch("core.ocr.get_ocr_provider", wraps=get_ocr_provider) as get_provider, patch(
                "core.ocr.factory.PaddleOCRProvider"
            ) as paddle:
                app = create_materials_app(root)
                client = app.test_client()
                workspace_id = self._new_workspace_id(client)
                self._set_valid_template(root, root / "workspaces" / workspace_id)
                client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"ocr_provider": "paddle", "receipts": (BytesIO(b"receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                )

            self.assertEqual(app.config["MATERIALS_OCR_PROVIDER"], "mock")
            self.assertTrue(app.config["MATERIALS_OFFLINE_OCR"])
            self.assertEqual(ingest.call_args.args[1], "mock")
            get_provider.assert_called_once_with("mock")
            paddle.assert_not_called()

    def test_wizard_steps_navigation_is_sticky(self):
        template = (Path("app") / "templates" / "materials" / "base.html").read_text(encoding="utf-8")
        self.assertRegex(
            template,
            r"\.steps\s*\{[^}]*position:sticky;[^}]*top:var\(--workspace-context-height\);[^}]*z-index:\d+;",
        )

    def test_receipt_action_column_stays_visible_during_horizontal_scroll(self):
        template = (Path("app") / "templates" / "materials" / "base.html").read_text(encoding="utf-8")

        self.assertRegex(
            template,
            r"\.receipt-table th:last-child,\.receipt-table td:last-child\s*\{[^}]*position:sticky;[^}]*right:0;",
        )

    def test_step_five_explains_strict_reconciled_autofill(self):
        template = (Path("app") / "templates" / "materials" / "wizard.html").read_text(encoding="utf-8")

        self.assertIn("数量×含税单价与价税合计一致", template)
        self.assertIn("其余字段留空供人工核对", template)

    def test_backup_download_and_restore_upload_open_a_new_workspace_at_step_one(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_materials_app(root)
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            (workspace_dir / "imports" / "receipt.pdf").write_bytes(b"receipt")

            backup = client.get(f"/workspace/{workspace_id}/backup")
            self.assertEqual(backup.status_code, 200)
            backup_data = backup.data
            backup.close()
            with ZipFile(BytesIO(backup_data)) as archive:
                self.assertIn("workspace.json", archive.namelist())
                self.assertIn("imports/receipt.pdf", archive.namelist())

            restored = client.post(
                "/workspace/restore",
                data={"backup": (BytesIO(backup_data), "workspace.zip")},
                content_type="multipart/form-data",
            )

            self.assertEqual(restored.status_code, 302)
            self.assertRegex(restored.headers["Location"], r"/workspace/[^/]+/step/1$")
            restored_id = restored.headers["Location"].split("/")[2]
            self.assertNotEqual(restored_id, workspace_id)
            restored_state = load_workspace(root / "workspaces" / restored_id)
            self.assertEqual(
                restored_state["template_dir"],
                str(Path(app.config["MATERIALS_TEMPLATE_DIR"]).resolve()),
            )
            self.assertEqual(restored_state["template_mode"], "bundled")

    def test_restore_upload_flashes_error_for_archive_without_workspace_file(self):
        with TemporaryDirectory() as directory:
            archive = BytesIO()
            with ZipFile(archive, "w") as backup:
                backup.writestr("imports/receipt.pdf", b"receipt")
            archive.seek(0)
            client = create_materials_app(directory).test_client()

            response = client.post(
                "/workspace/restore",
                data={"backup": (archive, "invalid.zip")},
                content_type="multipart/form-data",
                follow_redirects=True,
            )

            self.assertEqual(response.status_code, 200)
            self.assertIn("workspace.json", response.get_data(as_text=True))

    def test_invalid_template_directory_is_not_saved_and_returns_to_step_one(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_materials_app(root)
            app.config["MATERIALS_OCR_PROVIDER"] = "mock"
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            missing_templates = root / "missing-templates"

            response = client.post(
                f"/workspace/{workspace_id}/templates",
                data={"template_dir": str(missing_templates)},
            )

            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/step/1"))
            self.assertEqual(
                load_workspace(root / "workspaces" / workspace_id)["template_dir"],
                str(app.config["MATERIALS_TEMPLATE_DIR"]),
            )
            page = client.get(response.headers["Location"]).get_data(as_text=True)
            self.assertIn("模板无效", page)

    def test_template_directory_is_saved_as_absolute_path_and_survives_cwd_change(self):
        with TemporaryDirectory(dir=Path.cwd()) as directory:
            root = Path(directory)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单模板.xlsx")
            self._write_template(templates / "出库单模板.xlsx")
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            original_cwd = Path.cwd()
            relative_templates = os.path.relpath(templates, start=original_cwd)
            try:
                response = client.post(
                    f"/workspace/{workspace_id}/templates",
                    data={"template_dir": relative_templates},
                )
                saved = load_workspace(root / "workspaces" / workspace_id)
                os.chdir(root)
                step_response = client.get(f"/workspace/{workspace_id}/step/2")
            finally:
                os.chdir(original_cwd)

            self.assertEqual(response.status_code, 302)
            self.assertEqual(saved["template_dir"], str(templates.resolve()))
            self.assertEqual(step_response.status_code, 200)

    def test_direct_step_two_with_invalid_template_redirects_to_step_one(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            state = load_workspace(workspace_dir)
            state["template_dir"] = str(root / "missing-templates")
            save_workspace(workspace_dir, state)

            response = client.get(f"/workspace/{workspace_id}/step/2")

            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/step/1"))

    def test_step_three_shows_missing_receipts_and_next_action(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单模板.xlsx")
            self._write_template(templates / "出库单模板.xlsx")
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            client.post(
                f"/workspace/{workspace_id}/templates",
                data={"template_dir": str(templates)},
            )

            response = client.get(f"/workspace/{workspace_id}/step/3")

            self.assertEqual(response.status_code, 200)
            page = response.get_data(as_text=True)
            self.assertIn("未导入票据", page)
            self.assertIn("下一步", page)

    def test_step_four_and_quota_post_require_confirmed_receipts(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [{"receipt_id": "receipt-1", "is_material": True, "confirmed": False}]
            save_workspace(workspace_dir, state)

            get_response = client.get(f"/workspace/{workspace_id}/step/4")
            post_response = client.post(
                f"/workspace/{workspace_id}/quotas",
                data={"label": "Materials", "capacity": "100.00"},
            )

            self.assertEqual(get_response.status_code, 302)
            self.assertTrue(get_response.headers["Location"].endswith("/step/3"))
            self.assertEqual(post_response.status_code, 302)
            self.assertTrue(post_response.headers["Location"].endswith("/step/3"))
            page = client.get(post_response.headers["Location"]).get_data(as_text=True)
            self.assertIn("请先完成第 3 步", page)

    def test_open_export_rejects_traversal_and_uses_windows_startfile(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            export_dir = workspace_dir / "exports" / "ready-package"
            export_dir.mkdir()

            traversal = client.post(
                f"/workspace/{workspace_id}/exports/../imports/open"
            )
            with patch("materials_desktop.os.startfile") as startfile:
                response = client.post(
                    f"/workspace/{workspace_id}/exports/ready-package/open"
                )

            self.assertEqual(traversal.status_code, 404)
            self.assertEqual(response.status_code, 302)
            startfile.assert_called_once_with(str(export_dir.resolve()))

    def test_export_of_incomplete_package_shows_preflight_errors(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source = workspace_dir / "imports" / "receipt.png"
            source.write_bytes(b"receipt")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": str(source),
                "invoice_date": "2026-07-13", "total_cents": 10_000,
                "confirmed": True, "is_material": True,
            }]
            state["quotas"] = [{"slot_id": "quota-1", "capacity_cents": 10_000, "receipt_ids": []}]
            save_workspace(workspace_dir, state)
            slot_id = "quota-1"

            response = client.post(
                f"/workspace/{workspace_id}/export/{slot_id}"
            )

            self.assertEqual(response.status_code, 302)
            page = client.get(response.headers["Location"]).get_data(as_text=True)
            self.assertIn("导出前检查失败", page)

    def test_complete_http_flow_exports_material_package(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单模板.xlsx")
            self._write_template(templates / "出库单模板.xlsx")
            app = create_materials_app(root)
            app.config["MATERIALS_OCR_PROVIDER"] = "mock"
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)

            client.post(
                f"/workspace/{workspace_id}/templates",
                data={"template_dir": str(templates)},
            )
            client.post(
                f"/workspace/{workspace_id}/receipts/import",
                data={"receipts": (BytesIO(b"local receipt"), "receipt.pdf")},
                content_type="multipart/form-data",
            )
            state = load_workspace(root / "workspaces" / workspace_id)
            receipt_id = state["receipts"][0]["receipt_id"]
            client.post(
                f"/workspace/{workspace_id}/receipts/{receipt_id}",
                data={
                    "date": "2026-07-13",
                    "total": "100.00",
                    "vendor": "Local supplier",
                    "invoice": "INV-1",
                    "is_material": "on",
                },
            )
            client.post(
                f"/workspace/{workspace_id}/quotas",
                data={"label": "Materials", "capacity": "100.00"},
            )
            state = load_workspace(root / "workspaces" / workspace_id)
            slot_id = state["quotas"][0]["slot_id"]
            client.post(
                f"/workspace/{workspace_id}/quotas/{slot_id}/assign",
                data={"receipt_ids": [receipt_id]},
            )
            client.post(
                f"/workspace/{workspace_id}/quotas/{slot_id}/lines",
                data={
                    "inventory_date": "2026-07-13",
                    "name": "Sensor",
                    "specification": "S-1",
                    "unit": "piece",
                    "quantity": "2",
                    "unit_price": "50.00",
                    "amount": "100.00",
                    "supplier_name": "Local supplier",
                },
            )

            response = client.post(
                f"/workspace/{workspace_id}/export/{slot_id}", follow_redirects=True
            )

            self.assertIn("材料包已生成", response.get_data(as_text=True))
            self.assertIn("票据数量", response.get_data(as_text=True))
            state = load_workspace(root / "workspaces" / workspace_id)
            self.assertEqual(len(state["exports"]), 1)

    def test_quota_can_be_corrected_without_recreating_it(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "confirmed": True, "is_material": True,
                "total_cents": 10_000,
            }]
            state["quotas"] = [{
                "slot_id": "quota-1", "label": "错误名称",
                "capacity_cents": 20_000, "receipt_ids": ["receipt-1"],
            }]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/quotas/quota-1/update",
                data={"label": "项目一", "capacity": "150.00"},
            )

            self.assertEqual(response.status_code, 302)
            updated = load_workspace(workspace_dir)["quotas"][0]
            self.assertEqual(updated["label"], "项目一")
            self.assertEqual(updated["capacity_cents"], 15_000)
            page = client.get(f"/workspace/{workspace_id}/step/4").get_data(as_text=True)
            self.assertIn("更新额度", page)
            self.assertIn("删除额度", page)

    def test_quota_cannot_be_reduced_below_assigned_receipt_total(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "confirmed": True, "is_material": True,
                "total_cents": 10_000,
            }]
            state["quotas"] = [{
                "slot_id": "quota-1", "label": "项目一",
                "capacity_cents": 20_000, "receipt_ids": ["receipt-1"],
            }]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/quotas/quota-1/update",
                data={"label": "项目一", "capacity": "99.99"},
                follow_redirects=True,
            )

            self.assertIn("不能小于已分配票据合计", response.get_data(as_text=True))
            self.assertEqual(
                load_workspace(workspace_dir)["quotas"][0]["capacity_cents"], 20_000
            )

    def test_deleting_quota_keeps_receipts_and_removes_dependent_drafts(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "confirmed": True, "is_material": True,
                "total_cents": 10_000,
            }]
            state["quotas"] = [{
                "slot_id": "quota-1", "label": "项目一",
                "capacity_cents": 20_000, "receipt_ids": ["receipt-1"],
            }]
            state["lines_by_slot"] = {"quota-1": [{"name": "开发板"}]}
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/quotas/quota-1/delete"
            )

            self.assertEqual(response.status_code, 302)
            updated = load_workspace(workspace_dir)
            self.assertEqual(updated["quotas"], [])
            self.assertNotIn("quota-1", updated["lines_by_slot"])
            self.assertEqual(updated["receipts"][0]["receipt_id"], "receipt-1")

    def test_receipt_confirmation_rejects_zero_total(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [{"receipt_id": "receipt-1", "is_material": True}]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/receipt-1",
                data={"date": "2026-07-13", "total": "0", "is_material": "on"},
            )

            self.assertEqual(response.status_code, 302)
            self.assertFalse(load_workspace(workspace_dir)["receipts"][0].get("confirmed", False))
            page = client.get(response.headers["Location"]).get_data(as_text=True)
            self.assertIn("票据金额必须大于零", page)

    def test_saving_lines_keeps_multiple_rows_and_discards_blank_row(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source = workspace_dir / "imports" / "receipt.png"
            source.write_bytes(b"receipt")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": str(source), "invoice_date": "2026-07-13",
                "total_cents": 10_000, "confirmed": True, "is_material": True,
            }]
            state["quotas"] = [{"slot_id": "quota-1", "capacity_cents": 10_000, "receipt_ids": []}]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/quotas/quota-1/lines",
                data={
                    "inventory_date": ["2026-07-13", "2026-07-14", ""],
                    "name": ["Sensor", "Cable", ""],
                    "specification": ["S-1", "C-1", ""],
                    "unit": ["piece", "piece", ""],
                    "quantity": ["1", "2", ""],
                    "unit_price": ["50.00", "25.00", ""],
                    "amount": ["50.00", "50.00", ""],
                    "supplier_name": ["Vendor", "Vendor", ""],
                },
            )

            self.assertEqual(response.status_code, 302)
            lines = load_workspace(workspace_dir)["lines_by_slot"]["quota-1"]
            self.assertEqual([line["name"] for line in lines], ["Sensor", "Cable"])

    def test_auto_assign_uses_confirmed_receipts(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source = workspace_dir / "imports" / "receipt.png"
            source.write_bytes(b"receipt")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": str(source), "invoice_date": "2026-07-13",
                "total_cents": 10_000, "vendor_name": "Vendor", "ocr_text": "商品：C620 无刷电机调速器",
                "confirmed": True, "is_material": True,
            }]
            state["quotas"] = [{"slot_id": "quota-1", "capacity_cents": 10_000, "receipt_ids": []}]
            save_workspace(workspace_dir, state)

            response = client.post(f"/workspace/{workspace_id}/quotas/auto-assign")

            self.assertEqual(response.status_code, 302)
            updated_state = load_workspace(workspace_dir)
            self.assertEqual(updated_state["quotas"][0]["receipt_ids"], ["receipt-1"])
            self.assertEqual(updated_state["lines_by_slot"]["quota-1"], [{
                "inventory_date": "2026-07-13",
                "name": "",
                "specification": "",
                "unit": "",
                "quantity": "",
                "unit_price": "",
                "amount": "100.00",
                "supplier_name": "Vendor",
                "recognition_status": "待人工填写",
                "_auto_generated": True,
                "_source_receipt_id": "receipt-1",
            }])
            page = client.get(response.headers["Location"]).get_data(as_text=True)
            self.assertIn("260713_01_100.00.png", page)
            self.assertNotIn("receipt-1 (100.00)", page)

    def test_auto_assign_prefills_verified_line_items(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source = workspace_dir / "imports" / "receipt.pdf"
            source.write_bytes(b"receipt")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": str(source), "invoice_date": "2026-06-20",
                "total_cents": 96_600, "vendor_name": "Vendor", "confirmed": True, "is_material": True,
                "line_items": [{
                    "name": "RoboMaster C620无刷电机调速器", "specification": "C620", "unit": "件",
                    "quantity": "6", "unit_price_cents": 16_100, "amount_cents": 96_600,
                    "confidence": "verified",
                }],
            }]
            state["quotas"] = [{"slot_id": "quota-1", "capacity_cents": 100_000, "receipt_ids": []}]
            save_workspace(workspace_dir, state)

            response = client.post(f"/workspace/{workspace_id}/quotas/auto-assign")

            self.assertEqual(response.status_code, 302)
            line = load_workspace(workspace_dir)["lines_by_slot"]["quota-1"][0]
            self.assertEqual(
                (line["name"], line["specification"], line["unit"], line["quantity"], line["unit_price"], line["amount"]),
                ("RoboMaster C620无刷电机调速器", "C620", "件", "6", "161.00", "966.00"),
            )
            self.assertEqual(line["recognition_status"], "已校验")

    def test_auto_assign_replaces_known_bad_legacy_auto_drafts(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source = workspace_dir / "imports" / "receipt.pdf"
            source.write_bytes(b"receipt")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": str(source), "invoice_date": "2026-06-20",
                "total_cents": 96_600, "vendor_name": "Vendor",
                "ocr_text": "*导航遥控设备*RoboMast\ner C620 无刷电机调速器\n2件 6238.05",
                "confirmed": True, "is_material": True,
            }]
            state["quotas"] = [{"slot_id": "quota-1", "capacity_cents": 100_000, "receipt_ids": ["receipt-1"]}]
            state["lines_by_slot"] = {"quota-1": [{
                "inventory_date": "2026-06-20", "name": "电子发票 项目名称 规格型号 数量",
                "specification": "", "unit": "件", "quantity": "1", "unit_price": "966.00",
                "amount": "966.00", "supplier_name": "Vendor",
            }]}
            save_workspace(workspace_dir, state)

            response = client.post(f"/workspace/{workspace_id}/quotas/auto-assign")

            self.assertEqual(response.status_code, 302)
            line = load_workspace(workspace_dir)["lines_by_slot"]["quota-1"][0]
            self.assertEqual(line["name"], "")
            self.assertEqual(line["specification"], "")
            self.assertEqual(line["quantity"], "")
            self.assertEqual(line["unit_price"], "")
            self.assertEqual(line["recognition_status"], "待人工填写")

    def test_confirming_changed_amount_rebuilds_only_automatic_draft(self):
        with self._draft_refresh_workspace() as (client, workspace_id, workspace_dir):
            with patch("materials_desktop.verified_pdf_line_items", return_value=()):
                client.post(
                    f"/workspace/{workspace_id}/receipts/r1",
                    data={
                        "date": "2026-07-15", "total": "120.00",
                        "vendor": "供应商", "is_material": "on",
                    },
                )

            line = load_workspace(workspace_dir)["lines_by_slot"]["q1"][0]
            self.assertEqual(
                (line["name"], line["quantity"], line["unit_price"]),
                ("", "", ""),
            )
            self.assertEqual(
                (line["amount"], line["recognition_status"]),
                ("120.00", "待人工填写"),
            )

    def test_deleting_receipt_removes_its_automatic_draft(self):
        with self._draft_refresh_workspace() as (client, workspace_id, workspace_dir):
            client.post(f"/workspace/{workspace_id}/receipts/r1/delete")

            saved = load_workspace(workspace_dir)
            self.assertEqual(saved["quotas"][0]["receipt_ids"], [])
            self.assertNotIn("q1", saved["lines_by_slot"])

    def test_manual_assignment_change_regenerates_automatic_drafts(self):
        with self._draft_refresh_workspace() as (client, workspace_id, workspace_dir):
            client.post(
                f"/workspace/{workspace_id}/quotas/q1/assign",
                data={"receipt_ids": ["r2"]},
            )

            line = load_workspace(workspace_dir)["lines_by_slot"]["q1"][0]
            self.assertEqual(
                (line["_source_receipt_id"], line["amount"]),
                ("r2", "80.00"),
            )

    def test_receipt_change_never_overwrites_manual_inventory_lines(self):
        manual = {
            "inventory_date": "2026-07-15", "name": "人工名称",
            "specification": "M1", "unit": "件", "quantity": "2",
            "unit_price": "50.00", "amount": "100.00", "supplier_name": "供应商",
        }
        with self._draft_refresh_workspace() as (client, workspace_id, workspace_dir):
            state = load_workspace(workspace_dir)
            state["lines_by_slot"] = {"q1": [manual.copy()]}
            save_workspace(workspace_dir, state)

            with patch("materials_desktop.verified_pdf_line_items", return_value=()):
                client.post(
                    f"/workspace/{workspace_id}/receipts/r1",
                    data={
                        "date": "2026-07-15", "total": "120.00",
                        "vendor": "供应商", "is_material": "on",
                    },
                )

            self.assertEqual(load_workspace(workspace_dir)["lines_by_slot"]["q1"], [manual])

    def test_unchecking_material_removes_assignment_and_automatic_draft(self):
        with self._draft_refresh_workspace() as (client, workspace_id, workspace_dir):
            with patch("materials_desktop.verified_pdf_line_items", return_value=()):
                client.post(
                    f"/workspace/{workspace_id}/receipts/r1",
                    data={"date": "2026-07-15", "total": "100.00", "vendor": "供应商"},
                )

            saved = load_workspace(workspace_dir)
            self.assertEqual(saved["quotas"][0]["receipt_ids"], [])
            self.assertNotIn("q1", saved["lines_by_slot"])

    def test_auto_assign_refreshes_historical_selectable_pdf_items(self):
        with self._draft_refresh_workspace() as (client, workspace_id, workspace_dir):
            state = load_workspace(workspace_dir)
            state["receipts"] = [state["receipts"][0]]
            state["receipts"][0]["ocr_text"] = "electronic invoice text"
            state["receipts"][0]["line_items"] = []
            state["quotas"][0]["receipt_ids"] = []
            state["lines_by_slot"] = {}
            save_workspace(workspace_dir, state)
            verified = (
                RecognizedLineItem(
                    "RoboMaster C620 无刷电机调速器",
                    "C620",
                    "件",
                    "6",
                    16_100,
                    96_600,
                    "verified",
                ),
            )
            state = load_workspace(workspace_dir)
            state["receipts"][0]["total_cents"] = 96_600
            state["receipts"][0]["total"] = "966.00"
            state["quotas"][0]["capacity_cents"] = 100_000
            save_workspace(workspace_dir, state)

            with patch("materials_desktop.extract_pdf_text", return_value="electronic invoice"), patch(
                "materials_desktop.verified_pdf_line_items", return_value=verified
            ):
                client.post(f"/workspace/{workspace_id}/quotas/auto-assign")

            line = load_workspace(workspace_dir)["lines_by_slot"]["q1"][0]
            self.assertEqual(
                (line["name"], line["quantity"], line["unit_price"], line["amount"]),
                ("RoboMaster C620 无刷电机调速器", "6", "161.00", "966.00"),
            )

    def test_receipt_import_uses_configured_ocr_provider(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            bundle = root / "bundle"
            models = bundle / "ocr_models"
            self._write_required_ocr_models(models)
            with patch("materials_desktop.bundled_ocr_models_path", return_value=models), patch(
                "materials_desktop.bundle_root", return_value=bundle
            ), patch("materials_desktop.configure_bundled_ocr_assets"):
                app = create_materials_app(root)
            app.config["MATERIALS_OCR_PROVIDER"] = "test-provider"
            app.config["MATERIALS_STABLE_IMPORT"] = False
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            self._set_valid_template(root, root / "workspaces" / workspace_id)
            receipt = Receipt("receipt", "", None, None)

            with patch("materials_desktop.ingest_receipt", return_value=receipt) as ingest:
                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"receipts": (BytesIO(b"local receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                )

            self.assertEqual(response.status_code, 302)
            self.assertEqual(ingest.call_args.args[1], "test-provider")

    def test_scanned_pdf_import_uses_isolated_local_ocr_when_models_are_available(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            bundle = root / "bundle"
            models = bundle / "ocr_models"
            self._write_required_ocr_models(models)
            with patch(
                "materials_desktop.bundled_ocr_models_path", return_value=models
            ), patch(
                "materials_desktop.bundle_root", return_value=bundle
            ), patch("materials_desktop.configure_bundled_ocr_assets"):
                app = create_materials_app(root)
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            recognized = OCRResult(
                text=(
                    "开票日期：2026年06月20日 "
                    "销售方信息名称：深圳市睿炽科技有限公司 "
                    "价税合计（小写）¥966.00"
                ),
                provider="paddle",
            )

            with patch(
                "materials_desktop.recognize_with_worker",
                return_value=recognized,
            ) as recognize:
                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"receipts": (BytesIO(b"local receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                )

            saved = load_workspace(workspace_dir)["receipts"][0]
            self.assertEqual(response.status_code, 302)
            self.assertIn(f"/workspace/{workspace_id}/step/3", response.location)
            recognize.assert_called_once()
            self.assertEqual(saved["total_cents"], 96_600)
            self.assertEqual(saved["vendor_name"], "深圳市睿炽科技有限公司")
            self.assertIn("价税合计", saved["ocr_text"])

    def test_ocr_worker_failure_keeps_import_for_manual_confirmation(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            bundle = root / "bundle"
            models = bundle / "ocr_models"
            self._write_required_ocr_models(models)
            with patch(
                "materials_desktop.bundled_ocr_models_path", return_value=models
            ), patch(
                "materials_desktop.bundle_root", return_value=bundle
            ), patch("materials_desktop.configure_bundled_ocr_assets"):
                app = create_materials_app(root)
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)

            with patch(
                "materials_desktop.recognize_with_worker",
                return_value=OCRResult(
                    provider="paddle",
                    error="Local OCR worker exited with code 3221225477",
                ),
            ):
                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"receipts": (BytesIO(b"local receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                )

            saved = load_workspace(workspace_dir)["receipts"][0]
            self.assertEqual(response.status_code, 302)
            self.assertIn(f"/workspace/{workspace_id}/step/3", response.location)
            self.assertIsNone(saved["total_cents"])
            self.assertIn("exited with code", saved["extraction_note"])
            self.assertTrue(Path(saved["source_path"]).is_file())

    def test_receipt_import_preserves_original_chinese_pdf_suffix(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_materials_app(root)
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)

            with patch(
                "materials_desktop.ingest_receipt", return_value=Receipt("receipt", "", None, None)
            ) as ingest:
                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"receipts": (BytesIO(b"pdf receipt"), "发票.pdf")},
                    content_type="multipart/form-data",
                )

            saved_receipt = load_workspace(workspace_dir)["receipts"][0]
            source_path = Path(saved_receipt["source_path"])
            self.assertEqual(response.status_code, 302)
            self.assertEqual(Path(ingest.call_args.args[0]).suffix, ".pdf")
            self.assertEqual(source_path.suffix, ".pdf")
            self.assertTrue(source_path.is_file())

            preview = client.get(
                f"/workspace/{workspace_id}/receipts/{saved_receipt['receipt_id']}/source"
            )
            self.assertEqual(preview.status_code, 200)
            self.assertEqual(preview.mimetype, "application/pdf")
            self.assertNotIn("attachment", preview.headers.get("Content-Disposition", ""))
            preview.close()

    def test_receipt_import_rejects_unsupported_files_without_leaving_step_two(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            self._set_valid_template(root, root / "workspaces" / workspace_id)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/import",
                data={"receipts": (BytesIO(b"workbook"), "入库单.xlsx")},
                content_type="multipart/form-data",
                follow_redirects=True,
            )

            self.assertEqual(response.status_code, 200)
            self.assertTrue(response.request.path.endswith(f"/workspace/{workspace_id}/step/2"))
            self.assertIn("不支持的票据格式", response.get_data(as_text=True))
            self.assertEqual(load_workspace(root / "workspaces" / workspace_id)["receipts"], [])

    def test_receipt_import_requires_at_least_one_selected_file(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            self._set_valid_template(root, root / "workspaces" / workspace_id)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/import",
                data={},
                content_type="multipart/form-data",
                follow_redirects=True,
            )

            self.assertEqual(response.status_code, 200)
            self.assertTrue(response.request.path.endswith(f"/workspace/{workspace_id}/step/2"))
            self.assertIn("请至少选择一份票据", response.get_data(as_text=True))
            self.assertEqual(load_workspace(root / "workspaces" / workspace_id)["receipts"], [])

    def test_confirmation_table_uses_associated_forms_without_forms_inside_rows(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": "imports/receipt.png",
                "is_material": True,
            }]
            save_workspace(workspace_dir, state)

            page = client.get(f"/workspace/{workspace_id}/step/3").get_data(as_text=True)

            self.assertNotRegex(page, r"<tr>\s*<form")
            self.assertIn('form id="confirm-all-form"', page)
            self.assertIn("确认全部票据", page)
            self.assertIn("继续导入更多 PDF", page)
            self.assertIn("待确认", page)
            for field in ("date", "total", "vendor", "invoice", "is_material"):
                self.assertRegex(
                    page,
                    rf'<input[^>]*name="{field}-receipt-1"[^>]*form="confirm-all-form"',
                )

    def test_confirming_one_receipt_stays_on_step_three_and_reports_remaining_count(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [
                {"receipt_id": "receipt-1", "confirmed": False, "is_material": True},
                {"receipt_id": "receipt-2", "confirmed": False, "is_material": True},
            ]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/receipt-1",
                data={
                    "date-receipt-1": "2026-07-24",
                    "total-receipt-1": "12.34",
                    "vendor-receipt-1": "供应商",
                    "invoice-receipt-1": "INV-1",
                    "is_material-receipt-1": "on",
                },
                follow_redirects=True,
            )

            saved = load_workspace(workspace_dir)["receipts"]
            self.assertTrue(response.request.path.endswith(f"/workspace/{workspace_id}/step/3"))
            self.assertIn("还剩 1 张未确认", response.get_data(as_text=True))
            self.assertTrue(saved[0]["confirmed"])
            self.assertFalse(saved[1]["confirmed"])

    def test_confirm_all_receipts_updates_every_row_and_advances(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [
                {"receipt_id": "receipt-1", "confirmed": False, "is_material": True},
                {"receipt_id": "receipt-2", "confirmed": False, "is_material": True},
            ]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/confirm-all",
                data={
                    "date-receipt-1": "2026-07-24",
                    "total-receipt-1": "12.34",
                    "vendor-receipt-1": "供应商一",
                    "invoice-receipt-1": "INV-1",
                    "is_material-receipt-1": "on",
                    "date-receipt-2": "2026-07-23",
                    "total-receipt-2": "56.78",
                    "vendor-receipt-2": "供应商二",
                    "invoice-receipt-2": "INV-2",
                    "is_material-receipt-2": "on",
                },
                follow_redirects=True,
            )

            saved = load_workspace(workspace_dir)["receipts"]
            self.assertTrue(response.request.path.endswith(f"/workspace/{workspace_id}/step/4"))
            self.assertIn("已确认全部 2 张票据", response.get_data(as_text=True))
            self.assertTrue(all(receipt["confirmed"] for receipt in saved))
            self.assertEqual([receipt["total_cents"] for receipt in saved], [1234, 5678])

    def test_confirm_all_receipts_is_transactional_when_one_amount_is_invalid(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            state["receipts"] = [
                {"receipt_id": "receipt-1", "confirmed": False, "is_material": True},
                {"receipt_id": "receipt-2", "confirmed": False, "is_material": True},
            ]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/confirm-all",
                data={
                    "total-receipt-1": "12.34",
                    "is_material-receipt-1": "on",
                    "total-receipt-2": "",
                    "is_material-receipt-2": "on",
                },
                follow_redirects=True,
            )

            saved = load_workspace(workspace_dir)["receipts"]
            self.assertTrue(response.request.path.endswith(f"/workspace/{workspace_id}/step/3"))
            self.assertIn("第 2 张票据", response.get_data(as_text=True))
            self.assertFalse(any(receipt.get("confirmed", False) for receipt in saved))

    def test_separate_pdf_import_requests_append_instead_of_replacing(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)

            with patch(
                "materials_desktop.ingest_receipt",
                side_effect=[
                    Receipt("first", "", None, None),
                    Receipt("second", "", None, None),
                ],
            ):
                for filename in ("第一批.pdf", "第二批.pdf"):
                    response = client.post(
                        f"/workspace/{workspace_id}/receipts/import",
                        data={"receipts": (BytesIO(filename.encode()), filename)},
                        content_type="multipart/form-data",
                    )
                    self.assertEqual(response.status_code, 302)

            receipts = load_workspace(workspace_dir)["receipts"]
            self.assertEqual([receipt["original_filename"] for receipt in receipts], ["第一批.pdf", "第二批.pdf"])
            self.assertTrue(all(Path(receipt["source_path"]).is_file() for receipt in receipts))

    def test_image_receipt_import_is_rejected_with_pdf_only_guidance(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/import",
                data={"receipts": (BytesIO(b"payment screenshot"), "付款记录.png")},
                content_type="multipart/form-data",
                follow_redirects=True,
            )

            self.assertTrue(response.request.path.endswith(f"/workspace/{workspace_id}/step/2"))
            self.assertIn("仅支持 PDF", response.get_data(as_text=True))
            self.assertEqual(load_workspace(workspace_dir)["receipts"], [])

    def test_receipt_source_preview_is_scoped_to_workspace_imports(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source_path = workspace_dir / "imports" / "receipt.png"
            source_path.write_bytes(b"preview content")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1",
                "source_path": str(source_path),
                "is_material": True,
            }]
            save_workspace(workspace_dir, state)

            response = client.get(
                f"/workspace/{workspace_id}/receipts/receipt-1/source"
            )

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.data, b"preview content")
            self.assertTrue(response.mimetype.startswith("image/"))
            self.assertNotIn("attachment", response.headers.get("Content-Disposition", ""))
            response.close()

            download_path = workspace_dir / "imports" / "receipt.txt"
            download_path.write_bytes(b"download content")
            state["receipts"][0]["source_path"] = str(download_path)
            save_workspace(workspace_dir, state)
            response = client.get(
                f"/workspace/{workspace_id}/receipts/receipt-1/source"
            )
            self.assertEqual(response.status_code, 200)
            self.assertIn("attachment", response.headers.get("Content-Disposition", ""))
            response.close()

            outside_path = root / "outside.png"
            outside_path.write_bytes(b"outside")
            state["receipts"][0]["source_path"] = str(outside_path)
            save_workspace(workspace_dir, state)
            self.assertEqual(
                client.get(f"/workspace/{workspace_id}/receipts/receipt-1/source").status_code,
                404,
            )
            state["receipts"][0]["source_path"] = str(workspace_dir / "imports" / "missing.png")
            save_workspace(workspace_dir, state)
            self.assertEqual(
                client.get(f"/workspace/{workspace_id}/receipts/receipt-1/source").status_code,
                404,
            )

    def test_receipt_over_per_file_limit_is_rejected_transactionally(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_materials_app(root)
            app.config["MATERIALS_OFFLINE_OCR"] = True
            app.config["MAX_RECEIPT_BYTES"] = 10
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            self._set_valid_template(root, root / "workspaces" / workspace_id)

            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always", ResourceWarning)
                with BytesIO(b"x" * 11) as upload:
                    response = client.post(
                        f"/workspace/{workspace_id}/receipts/import",
                        data={"receipts": (upload, "large.pdf")},
                        follow_redirects=True,
                    )
                try:
                    self.assertEqual(response.status_code, 200)
                    self.assertTrue(response.request.path.endswith(f"/workspace/{workspace_id}/step/2"))
                    self.assertIn("单个票据不能超过", response.get_data(as_text=True))
                    self.assertEqual(
                        load_workspace(root / "workspaces" / workspace_id)["receipts"], []
                    )
                finally:
                    response.close()
                gc.collect()

            self.assertFalse([warning for warning in caught if warning.category is ResourceWarning])

    def test_batch_upload_accepts_multiple_receipts_within_each_file_limit(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_materials_app(root)
            app.config["MATERIALS_OFFLINE_OCR"] = True
            self.assertGreater(
                app.config["MAX_CONTENT_LENGTH"],
                2 * app.config["MAX_RECEIPT_BYTES"],
            )
            app.config["MAX_RECEIPT_BYTES"] = 10
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            self._set_valid_template(root, root / "workspaces" / workspace_id)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/import",
                data={
                    "receipts": [
                        (BytesIO(b"x" * 8), "first.pdf"),
                        (BytesIO(b"y" * 8), "second.pdf"),
                    ]
                },
                content_type="multipart/form-data",
            )

            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/step/3"))
            self.assertEqual(
                len(load_workspace(root / "workspaces" / workspace_id)["receipts"]), 2
            )

    def test_index_migrates_legacy_root_workspace_and_keeps_imports_accessible(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            legacy_import = root / "imports" / "legacy-receipt.png"
            legacy_import.parent.mkdir()
            legacy_import.write_bytes(b"legacy receipt")
            (root / "workspace.json").write_text(
                '{"name":"Legacy project","template_dir":"legacy-templates",'
                '"ocr_provider":"mock","receipts":[{"receipt_id":"legacy-1",'
                f'"source_path":"{legacy_import.as_posix()}","is_material":true}}],'
                '"quotas":[]}',
                encoding="utf-8",
            )

            client = create_materials_app(root).test_client()
            page = client.get("/").get_data(as_text=True)
            migrated = list((root / "workspaces").iterdir())

            self.assertIn("legacy-", page)
            self.assertEqual(len(migrated), 1)
            self.assertTrue(migrated[0].name.startswith("legacy-"))
            state = load_workspace(migrated[0])
            self.assertEqual(state["name"], "Legacy project")
            self.assertEqual(
                Path(state["receipts"][0]["source_path"]).read_bytes(), b"legacy receipt"
            )

            create_materials_app(root)
            self.assertEqual(len(list((root / "workspaces").iterdir())), 1)

    def test_receipt_import_rolls_back_when_a_later_upload_fails(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_materials_app(root)
            app.config["MATERIALS_OCR_PROVIDER"] = "mock"
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)

            save_calls = 0

            def save_first_upload_then_fail(_upload, destination, *_args, **_kwargs):
                nonlocal save_calls
                save_calls += 1
                if save_calls == 2:
                    raise OSError("disk full")
                Path(destination).write_bytes(b"first")

            with patch(
                "werkzeug.datastructures.FileStorage.save",
                autospec=True,
                side_effect=save_first_upload_then_fail,
            ):
                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={
                        "receipts": [
                            (BytesIO(b"first"), "first.pdf"),
                            (BytesIO(b"second"), "second.pdf"),
                        ],
                    },
                    content_type="multipart/form-data",
                )

            self.assertEqual(response.status_code, 302)
            self.assertEqual(load_workspace(workspace_dir)["receipts"], [])
            self.assertEqual(list((workspace_dir / "imports").iterdir()), [])
            self.assertFalse((workspace_dir / ".import-staging").exists())

    def test_receipt_import_rolls_back_when_workspace_save_fails(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_materials_app(root)
            app.config["MATERIALS_OCR_PROVIDER"] = "mock"
            client = app.test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)

            with patch("materials_desktop.save_workspace", side_effect=OSError("disk full")):
                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"receipts": (BytesIO(b"receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                )

            self.assertEqual(response.status_code, 302)
            self.assertEqual(load_workspace(workspace_dir)["receipts"], [])
            self.assertEqual(list((workspace_dir / "imports").iterdir()), [])
            self.assertFalse((workspace_dir / ".import-staging").exists())

    def test_quota_save_oserror_redirects_with_actionable_flash(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source = workspace_dir / "imports" / "receipt.png"
            source.write_bytes(b"receipt")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{"receipt_id": "receipt-1", "source_path": str(source), "is_material": True, "confirmed": True}]
            save_workspace(workspace_dir, state)

            with patch("materials_desktop.save_workspace", side_effect=OSError("disk full")):
                response = client.post(
                    f"/workspace/{workspace_id}/quotas",
                    data={"label": "Materials", "capacity": "100.00"},
                )

            self.assertEqual(response.status_code, 302)
            page = client.get(response.headers["Location"]).get_data(as_text=True)
            self.assertIn("本地文件写入失败", page)

    def test_upload_save_oserror_does_not_persist_receipt_record(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)

            with patch(
                "werkzeug.datastructures.file_storage.FileStorage.save",
                side_effect=OSError("disk full"),
            ):
                response = client.post(
                    f"/workspace/{workspace_id}/receipts/import",
                    data={"receipts": (BytesIO(b"local receipt"), "receipt.pdf")},
                    content_type="multipart/form-data",
                )

            self.assertEqual(response.status_code, 302)
            self.assertEqual(load_workspace(workspace_dir)["receipts"], [])
            page = client.get(response.headers["Location"]).get_data(as_text=True)
            self.assertIn("本地文件写入失败", page)

    def test_every_post_route_redirects_with_storage_error(self):
        routes = (
            "new_workspace", "templates", "import_receipts", "confirm_receipt", "confirm_all_receipts", "add_quota",
            "assign_receipts", "auto_assign_receipts", "save_lines", "delete_receipt", "export_package",
        )
        for route in routes:
            with self.subTest(route=route), TemporaryDirectory() as directory:
                root = Path(directory)
                app = create_materials_app(root)
                app.config["MATERIALS_OCR_PROVIDER"] = "mock"
                client = app.test_client()
                if route == "new_workspace":
                    with patch("materials_desktop.create_workspace", side_effect=OSError("disk full")):
                        response = client.post("/workspace/new", data={"name": "Project Alpha"})
                else:
                    workspace_id = self._new_workspace_id(client)
                    workspace_dir = root / "workspaces" / workspace_id
                    templates = root / "templates"
                    templates.mkdir()
                    self._write_template(templates / "入库单模板.xlsx")
                    self._write_template(templates / "出库单模板.xlsx")
                    source = workspace_dir / "imports" / "receipt.png"
                    source.write_bytes(b"receipt")
                    state = load_workspace(workspace_dir)
                    state["receipts"] = [{
                        "receipt_id": "receipt-1", "source_path": str(source),
                        "invoice_date": "2026-07-13", "total_cents": 10_000,
                        "confirmed": True, "is_material": True,
                    }]
                    state["quotas"] = [{
                        "slot_id": "quota-1", "capacity_cents": 10_000,
                        "receipt_ids": ["receipt-1"],
                    }]
                    state["lines_by_slot"] = {"quota-1": [{
                        "inventory_date": "2026-07-13", "name": "Sensor",
                        "specification": "S-1", "unit": "piece", "quantity": "2",
                        "unit_price": "50.00", "amount": "100.00", "supplier_name": "Vendor",
                    }]}
                    state["template_dir"] = str(templates)
                    save_workspace(workspace_dir, state)
                    route_data = {
                        "templates": (
                            f"/workspace/{workspace_id}/templates",
                            {"template_dir": str(templates)},
                        ),
                        "import_receipts": (
                            f"/workspace/{workspace_id}/receipts/import",
                            {"receipts": (BytesIO(b"receipt"), "receipt.pdf")},
                        ),
                        "confirm_receipt": (
                            f"/workspace/{workspace_id}/receipts/receipt-1",
                            {"date": "2026-07-13", "total": "100.00", "is_material": "on"},
                        ),
                        "confirm_all_receipts": (
                            f"/workspace/{workspace_id}/receipts/confirm-all",
                            {
                                "date-receipt-1": "2026-07-13",
                                "total-receipt-1": "100.00",
                                "is_material-receipt-1": "on",
                            },
                        ),
                        "delete_receipt": (
                            f"/workspace/{workspace_id}/receipts/receipt-1/delete", {},
                        ),
                        "add_quota": (
                            f"/workspace/{workspace_id}/quotas",
                            {"label": "More", "capacity": "100.00"},
                        ),
                        "assign_receipts": (
                            f"/workspace/{workspace_id}/quotas/quota-1/assign",
                            {"receipt_ids": ["receipt-1"]},
                        ),
                        "auto_assign_receipts": (f"/workspace/{workspace_id}/quotas/auto-assign", {}),
                        "save_lines": (
                            f"/workspace/{workspace_id}/quotas/quota-1/lines",
                            {
                                "inventory_date": "2026-07-13", "name": "Sensor",
                                "specification": "S-1", "unit": "piece", "quantity": "2",
                                "unit_price": "50.00", "amount": "100.00", "supplier_name": "Vendor",
                            },
                        ),
                    }
                    if route == "export_package":
                        with patch(
                            "materials_desktop.export_quota_package",
                            side_effect=OSError("disk full"),
                        ):
                            response = client.post(f"/workspace/{workspace_id}/export/quota-1")
                    else:
                        target, data = route_data[route]
                        kwargs = {"data": data}
                        if route == "import_receipts":
                            kwargs["content_type"] = "multipart/form-data"
                        with patch("materials_desktop.save_workspace", side_effect=OSError("disk full")):
                            response = client.post(target, **kwargs)

                self.assertEqual(response.status_code, 302)
                page = client.get(response.headers["Location"]).get_data(as_text=True)
                self.assertRegex(page, "磁盘空间|权限")

    def test_confirmation_page_hides_absolute_paths_and_shows_ocr_text(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            state = load_workspace(workspace_dir)
            source_path = workspace_dir / "imports" / "receipt.png"
            source_path.write_bytes(b"receipt")
            state["receipts"] = [{
                "receipt_id": "receipt-1",
                "source_path": str(source_path),
                "ocr_text": "OCR sample text",
                "is_material": True,
            }]
            save_workspace(workspace_dir, state)

            page = client.get(f"/workspace/{workspace_id}/step/3").get_data(as_text=True)

            self.assertNotIn(str(source_path), page)
            self.assertIn("receipt.png", page)
            self.assertIn("OCR sample text", page)
            self.assertIn(
                f'href="/workspace/{workspace_id}/receipts/receipt-1/source"', page
            )
            self.assertIn('class="receipt-table"', page)
            self.assertIn(
                f'action="/workspace/{workspace_id}/receipts/receipt-1/delete"', page
            )
            self.assertIn("确认删除这张票据吗？", page)

    def test_delete_receipt_removes_workspace_file_and_quota_assignment(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            client = create_materials_app(root).test_client()
            workspace_id = self._new_workspace_id(client)
            workspace_dir = root / "workspaces" / workspace_id
            self._set_valid_template(root, workspace_dir)
            source_path = workspace_dir / "imports" / "wrong-receipt.png"
            source_path.write_bytes(b"receipt")
            state = load_workspace(workspace_dir)
            state["receipts"] = [{
                "receipt_id": "receipt-1", "source_path": str(source_path),
                "confirmed": True, "is_material": True,
            }]
            state["quotas"] = [{
                "slot_id": "quota-1", "capacity_cents": 10_000,
                "receipt_ids": ["receipt-1"],
            }]
            save_workspace(workspace_dir, state)

            response = client.post(
                f"/workspace/{workspace_id}/receipts/receipt-1/delete"
            )

            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/step/3"))
            saved = load_workspace(workspace_dir)
            self.assertEqual(saved["receipts"], [])
            self.assertEqual(saved["quotas"][0]["receipt_ids"], [])
            self.assertFalse(source_path.exists())

    def test_wizard_header_shows_workspace_name_but_index_does_not(self):
        with TemporaryDirectory() as directory:
            client = create_materials_app(directory).test_client()
            workspace_id = self._new_workspace_id(client)

            index = client.get("/").get_data(as_text=True)
            wizard = client.get(f"/workspace/{workspace_id}/step/1").get_data(as_text=True)

            self.assertNotIn("当前工作区：", index)
            self.assertIn("当前工作区：Project Alpha", wizard)

    @staticmethod
    def _set_valid_template(root: Path, workspace_dir: Path) -> Path:
        templates = root / "templates"
        templates.mkdir(exist_ok=True)
        MaterialsDesktopTests._write_template(templates / "入库单模板.xlsx")
        MaterialsDesktopTests._write_template(templates / "出库单模板.xlsx")
        state = load_workspace(workspace_dir)
        state["template_dir"] = str(templates)
        save_workspace(workspace_dir, state)
        return templates

    @staticmethod
    def _new_workspace_id(client) -> str:
        response = client.post("/workspace/new", data={"name": "Project Alpha"})
        return response.headers["Location"].split("/")[2]

    @staticmethod
    def _write_template(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        headers = [
            "入库日期" if "入库单" in path.stem else "出库日期",
            "产品名称",
            "规格型号",
            "单位",
            "入库数量" if "入库单" in path.stem else "出库数量",
            "单价(元)",
            "金额(元)",
            "供货单位" if "入库单" in path.stem else "用途",
            "经办人" if "入库单" in path.stem else "领用人",
            "管理员",
        ]
        for column, header in enumerate(headers, start=1):
            sheet.cell(4, column).value = header
        sheet.cell(7, 6).value = "合计："
        sheet.cell(7, 7).value = "=SUM(G5:G6)"
        sheet.cell(7, 8).value = "大写："
        sheet.cell(7, 9).value = '=TEXT(G7,"[DBNUM2]")&"元整"'
        try:
            workbook.save(path)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
