from datetime import date
from pathlib import Path
from random import Random
from tempfile import TemporaryDirectory
from types import SimpleNamespace
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from core.ocr import OCRResult

from core.materials_assistant import (
    InventoryLine,
    QuotaSlot,
    Receipt,
    allocate_receipts_to_quota_slots,
    amount_bucket,
    build_inventory_line_drafts,
    build_readable_filename,
    copy_renamed_receipts,
    find_candidate_packages,
    extract_pdf_text_boxes,
    ingest_receipt,
    parse_receipt_text,
    render_inventory_documents,
    render_receipt_markdown,
    validate_inventory_lines,
    validate_package_receipts,
    verified_pdf_line_items,
)
from core.inventory_line_extraction import TextBox
from core.inventory_line_extraction import RecognizedLineItem


class MaterialsAssistantTests(unittest.TestCase):
    def test_uses_largest_currency_summary_for_tax_inclusive_total(self):
        receipt = parse_receipt_text(
            "发票号码：26432000000045328426 "
            "开票日期：2026年01月07日 "
            "价税合计（小写）¥2299.00 "
            "合计金额¥2276.24 税额¥22.76"
        )

        self.assertEqual(receipt.total_cents, 229_900)

    def test_prefers_explicit_tax_inclusive_total_for_discounted_invoice(self):
        receipt = parse_receipt_text(
            "合计 ¥1000.00 "
            "税额 ¥130.00 "
            "价税合计（小写）¥900.00"
        )

        self.assertEqual(receipt.total_cents, 90_000)

    def test_tax_total_label_does_not_capture_later_invoice_number(self):
        receipt = parse_receipt_text(
            "合计 价税合计（大写）（小写）备注 开票人："
            "25432000000164978653 2025年12月03日 "
            "销售方 ¥184.85 税额 ¥1.85 ¥186.70"
        )

        self.assertEqual(receipt.total_cents, 18_670)

    def test_extracts_seller_names_for_small_business_suffixes(self):
        cases = (
            ("长沙市芙蓉区就好电子产品商行（个体工商户）", "92430102MADG5A9D3F"),
            ("吉首市鑫锐电脑经营部", "92433101MA4LR1LJ23"),
        )
        for seller, seller_tax_id in cases:
            with self.subTest(seller=seller):
                receipt = parse_receipt_text(
                    "名称：湘潭大学 "
                    "统一社会信用代码/纳税人识别号：124300004448750350 "
                    f"名称：{seller} "
                    f"统一社会信用代码/纳税人识别号：{seller_tax_id}"
                )

                self.assertEqual(receipt.vendor_name, seller)

    def test_verified_pdf_line_items_uses_the_confirmed_total(self):
        boxes = [
            TextBox("C620 无刷电机调速器", 10, 100), TextBox("C620", 260, 100),
            TextBox("件", 350, 100), TextBox("6", 430, 100),
            TextBox("161.00", 510, 100), TextBox("966.00", 610, 100),
        ]
        with patch("core.materials_assistant.extract_pdf_text_boxes", return_value=boxes):
            verified = verified_pdf_line_items("invoice.pdf", 96_600)
            conflicting = verified_pdf_line_items("invoice.pdf", 95_000)

        self.assertEqual(verified[0].confidence, "verified")
        self.assertEqual(conflicting, ())

    def test_pdf_text_boxes_combine_page_and_text_matrices(self):
        class FakePage:
            mediabox = SimpleNamespace(height=297.0)

            def extract_text(self, visitor_text):
                visitor_text(
                    "RoboMaster C620",
                    [1, 0, 0, -1, 12.755, 238.649],
                    [1, 0, 0, -1, 0, 19.435],
                    None,
                    10,
                )

        reader = SimpleNamespace(pages=[FakePage()])
        with patch(
            "core.materials_assistant._pdf_reader_types",
            return_value=(lambda _path: reader, ()),
        ):
            boxes = extract_pdf_text_boxes("invoice.pdf")

        self.assertAlmostEqual(boxes[0].x, 12.755, places=3)
        self.assertAlmostEqual(boxes[0].y, 77.786, places=3)

    def test_parse_electronic_invoice_text_and_render_markdown(self):
        receipt = parse_receipt_text(
            "电子发票（普通发票） 发票号码: 26957000000000722043 "
            "开票日期: 2026年06月20日 销售方信息名称：深圳市睿炽科技有限公司 "
            "价税合计（大写）玖佰陆拾陆圆整（小写）¥966.00",
            source_path="C620.pdf",
        )

        self.assertEqual(receipt.invoice_number, "26957000000000722043")
        self.assertEqual(receipt.invoice_date, date(2026, 6, 20))
        self.assertEqual(receipt.vendor_name, "深圳市睿炽科技有限公司")
        self.assertEqual(receipt.total_cents, 96600)
        self.assertEqual(amount_bucket(receipt.total_cents), "500-1000")
        self.assertEqual(build_readable_filename(receipt, 1), "260620_01_966.00.pdf")

        markdown = render_receipt_markdown(receipt)
        self.assertIn("价税合计: 966.00", markdown)
        self.assertIn("深圳市睿炽科技有限公司", markdown)

    def test_parse_layout_split_electronic_invoice_fields_conservatively(self):
        receipt = parse_receipt_text(
            "电子发票发票号码：开票日期：购买方信息统一社会信用代码/纳税人识别号："
            "销售方信息统一社会信用代码/纳税人识别号：名称：名称：项目名称"
            "263720000018386078262026年04月22日湘潭大学124300004448750350"
            "临沂市国金护理用品有限公司91371301MA3RYU9R2K¥168.51¥1.69¥170.20",
            source_path="fan.pdf",
        )

        self.assertEqual(receipt.invoice_number, "26372000001838607826")
        self.assertEqual(receipt.vendor_name, "临沂市国金护理用品有限公司")
        self.assertEqual(receipt.total_cents, 17020)

    def test_candidate_packages_respect_cap_and_keep_receipts_whole(self):
        receipts = [
            Receipt("a", "a.pdf", date(2026, 7, 1), 80000),
            Receipt("b", "b.pdf", date(2026, 7, 2), 70000),
            Receipt("c", "c.pdf", date(2026, 7, 3), 60000),
            Receipt("d", "d.pdf", date(2026, 7, 4), 10000),
        ]

        candidates = find_candidate_packages(receipts, 150000, limit=3)

        self.assertEqual(candidates[0].total_cents, 150000)
        self.assertEqual(set(candidates[0].receipt_ids), {"a", "b"})
        self.assertTrue(all(candidate.total_cents <= 150000 for candidate in candidates))
        self.assertTrue(all(len(candidate.receipt_ids) == len(set(candidate.receipt_ids)) for candidate in candidates))

        excluded = Receipt("travel", "travel.pdf", date(2026, 7, 3), 140000, is_material=False)
        self.assertEqual(find_candidate_packages([excluded], 150000), [])

    def test_receipt_copy_uses_readable_date_sequence_amount_name(self):
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            source = root / "原始票据.pdf"
            source.write_bytes(b"receipt")
            receipt = Receipt("receipt-1", str(source), date(2026, 7, 11), 10088)

            copied = copy_renamed_receipts([receipt], root / "原始票据")

            self.assertEqual(copied[0].name, "260711_01_100.88.pdf")
            self.assertEqual(copied[0].read_bytes(), b"receipt")

    def test_inventory_draft_leaves_ambiguous_ocr_item_fields_blank(self):
        receipt = Receipt(
            "receipt-1",
            "receipt.png",
            date(2026, 7, 13),
            10_000,
            vendor_name="Vendor",
            ocr_text="商品：C620 无刷电机调速器",
        )

        self.assertEqual(build_inventory_line_drafts([receipt]), [{
            "inventory_date": "2026-07-13",
            "name": "",
            "specification": "",
            "unit": "",
            "quantity": "",
            "unit_price": "",
            "amount": "100.00",
            "supplier_name": "Vendor",
            "recognition_status": "待人工填写",
            "_auto_generated": True,
            "_source_receipt_id": "receipt-1",
        }])

    def test_inventory_draft_does_not_use_unverified_invoice_text(self):
        receipt = Receipt(
            "receipt-2",
            "receipt.pdf",
            date(2026, 6, 20),
            96_600,
            vendor_name="Vendor",
            ocr_text=(
                "项目名称 规格型号 单位 数量 单价 金额\n"
                "*导航遥控设备*RoboMast\ner C620 无刷电机调速器\n"
                "2件 6238.05 13% 185.68"
            ),
        )

        self.assertEqual(build_inventory_line_drafts([receipt]), [{
            "inventory_date": "2026-06-20",
            "name": "",
            "specification": "",
            "unit": "",
            "quantity": "",
            "unit_price": "",
            "amount": "966.00",
            "supplier_name": "Vendor",
            "recognition_status": "待人工填写",
            "_auto_generated": True,
            "_source_receipt_id": "receipt-2",
        }])

    def test_inventory_draft_prefills_a_verified_line_item(self):
        receipt = Receipt(
            "receipt-3",
            "receipt.pdf",
            date(2026, 6, 20),
            96_600,
            vendor_name="Vendor",
            line_items=(
                RecognizedLineItem(
                    "RoboMaster C620无刷电机调速器", "C620", "件", "6", 16_100, 96_600, "verified"
                ),
            ),
        )

        self.assertEqual(build_inventory_line_drafts([receipt]), [{
            "inventory_date": "2026-06-20",
            "name": "RoboMaster C620无刷电机调速器",
            "specification": "C620",
            "unit": "件",
            "quantity": "6",
            "unit_price": "161.00",
            "amount": "966.00",
            "supplier_name": "Vendor",
            "recognition_status": "已校验",
            "_auto_generated": True,
            "_source_receipt_id": "receipt-3",
        }])

    def test_inventory_draft_ignores_stored_pending_line_items(self):
        receipt = Receipt(
            "receipt-pending",
            "receipt.pdf",
            date(2026, 4, 22),
            17_020,
            vendor_name="Vendor",
            line_items=(
                RecognizedLineItem(
                    "错误名称", "错误规格", "件", "168.51", 16_851, 169, "pending"
                ),
            ),
        )

        draft = build_inventory_line_drafts([receipt])[0]

        self.assertEqual(
            [draft[key] for key in ("name", "specification", "unit", "quantity", "unit_price")],
            ["", "", "", "", ""],
        )
        self.assertEqual(draft["amount"], "170.20")
        self.assertEqual(draft["recognition_status"], "待人工填写")
        self.assertTrue(draft["_auto_generated"])

    def test_material_and_source_receipt_totals_must_match(self):
        line = InventoryLine(date(2026, 7, 11), "开发板", "F4", "块", "2", 5000, 10002)
        with self.assertRaises(ValueError):
            validate_inventory_lines([line])

        receipt = Receipt("source", "source.pdf", date(2026, 7, 11), 10000)
        self.assertEqual(validate_package_receipts([receipt]), 10000)
        with self.assertRaises(ValueError):
            validate_package_receipts([receipt, receipt])

    def test_inventory_line_allows_one_cent_unit_price_rounding(self):
        line = InventoryLine(
            date(2026, 1, 7),
            "电脑 超小型计算机产品",
            "NUC11PHI7C",
            "台",
            "3",
            76_633,
            229_900,
        )

        self.assertEqual(validate_inventory_lines([line], expected_total_cents=229_900), 229_900)

    def test_inventory_line_rejects_non_finite_quantities_as_value_errors(self):
        for quantity in ("NaN", "Infinity", "1e999999"):
            with self.subTest(quantity=quantity):
                line = InventoryLine(
                    date(2026, 7, 11), "开发板", "F4", "块", quantity, 10_000, 10_000
                )
                with self.assertRaises(ValueError):
                    validate_inventory_lines([line])

    def test_confirmed_quota_slots_allocate_without_overflow_or_reuse(self):
        receipts = [
            Receipt("r-1200", "1200.pdf", date(2026, 7, 1), 120000),
            Receipt("r-900", "900.pdf", date(2026, 7, 2), 90000),
            Receipt("r-600", "600.pdf", date(2026, 7, 3), 60000),
            Receipt("r-200", "200.pdf", date(2026, 7, 4), 20000),
            Receipt("travel", "travel.pdf", date(2026, 7, 5), 10000, is_material=False),
        ]
        slots = [QuotaSlot("1500-01", 150000), QuotaSlot("2000-01", 200000)]

        allocation = allocate_receipts_to_quota_slots(receipts, slots)

        self.assertEqual(allocation.allocated_total_cents, 290000)
        self.assertEqual(allocation.unassigned_receipt_ids, ("travel",))
        allocated_ids = [
            receipt_id
            for package in allocation.packages
            for receipt_id in package.receipt_ids
        ]
        self.assertEqual(len(allocated_ids), len(set(allocated_ids)))
        self.assertTrue(all(package.total_cents <= package.capacity_cents for package in allocation.packages))

    def test_large_receipt_batch_still_finds_the_optimal_quota_allocation(self):
        amounts = (52, 36, 15, 48, 33, 15, 49, 57, 37, 53, 32, 36, 59, 48, 23, 22, 18)
        receipts = [
            Receipt(f"r-{index:02d}", f"{index:02d}.pdf", date(2026, 7, 1), amount)
            for index, amount in enumerate(amounts)
        ]
        slots = [QuotaSlot("q-1", 100), QuotaSlot("q-2", 100)]

        allocation = allocate_receipts_to_quota_slots(receipts, slots)

        self.assertEqual(allocation.allocated_total_cents, 200)
        self.assertTrue(all(package.total_cents <= 100 for package in allocation.packages))

    def test_typical_forty_receipt_batch_can_fill_all_project_quotas(self):
        random = Random(20260723)
        receipts = [
            Receipt(
                f"r-{index:02d}",
                f"{index:02d}.pdf",
                date(2026, 7, 1),
                random.randrange(5_000, 100_001, 100),
            )
            for index in range(40)
        ]
        slots = [
            QuotaSlot("q-1500", 150_000),
            QuotaSlot("q-2000", 200_000),
            QuotaSlot("q-6000", 600_000),
            QuotaSlot("q-8000", 800_000),
        ]

        allocation = allocate_receipts_to_quota_slots(receipts, slots)

        self.assertEqual(
            [package.total_cents for package in allocation.packages],
            [slot.capacity_cents for slot in slots],
        )

    def test_mock_ocr_never_creates_fake_financial_data_for_real_receipts(self):
        receipt = ingest_receipt("unreadable-scan.jpg", provider_name="mock")

        self.assertIsNone(receipt.invoice_date)
        self.assertIsNone(receipt.total_cents)
        self.assertIn("mock OCR", receipt.extraction_note)

    def test_ingest_image_uses_injected_local_recognizer(self):
        recognized_paths = []

        def recognize(path):
            recognized_paths.append(path)
            return OCRResult(
                text=(
                    "开票日期：2026年06月20日 "
                    "销售方信息名称：深圳市睿炽科技有限公司 "
                    "价税合计（小写）¥966.00"
                ),
                provider="paddle",
            )

        receipt = ingest_receipt(
            "scan.png",
            provider_name="paddle",
            recognizer=recognize,
        )

        self.assertEqual(recognized_paths, ["scan.png"])
        self.assertEqual(receipt.invoice_date, date(2026, 6, 20))
        self.assertEqual(receipt.total_cents, 96_600)
        self.assertEqual(receipt.vendor_name, "深圳市睿炽科技有限公司")

    def test_ingest_image_preserves_worker_error_for_manual_confirmation(self):
        receipt = ingest_receipt(
            "scan.png",
            provider_name="paddle",
            recognizer=lambda _path: OCRResult(
                provider="paddle",
                error="Local OCR worker exited with code 3221225477",
            ),
        )

        self.assertIsNone(receipt.total_cents)
        self.assertIn("exited with code", receipt.extraction_note)

    def test_ingest_electronic_pdf_does_not_call_injected_recognizer(self):
        parsed = Receipt("invoice", "invoice.pdf", date(2026, 6, 20), 96_600)

        def recognize(_path):
            raise AssertionError("Electronic PDFs must not use OCR")

        with patch(
            "core.materials_assistant.extract_pdf_text",
            return_value="electronic invoice",
        ), patch(
            "core.materials_assistant.parse_receipt_text",
            return_value=parsed,
        ), patch(
            "core.materials_assistant.verified_pdf_line_items",
            return_value=(),
        ):
            receipt = ingest_receipt(
                "invoice.pdf",
                provider_name="paddle",
                recognizer=recognize,
            )

        self.assertEqual(receipt.total_cents, 96_600)

    def test_ingest_receipt_keeps_verified_line_items_from_pdf_positions(self):
        parsed = Receipt("invoice", "invoice.pdf", date(2026, 6, 20), 96_600)
        boxes = [
            TextBox("*导航遥控设备*RoboMaster C620无刷电机调速器", 10, 100),
            TextBox("C620", 260, 100), TextBox("件", 350, 100),
            TextBox("6", 430, 100), TextBox("161.00", 510, 100),
            TextBox("966.00", 610, 100),
        ]

        with patch("core.materials_assistant.extract_pdf_text", return_value="electronic invoice"), patch(
            "core.materials_assistant.extract_pdf_text_boxes", return_value=boxes
        ), patch("core.materials_assistant.parse_receipt_text", return_value=parsed):
            receipt = ingest_receipt("invoice.pdf", provider_name="mock")

        self.assertEqual(receipt.line_items[0].quantity, "6")
        self.assertEqual(receipt.line_items[0].unit_price_cents, 16_100)
        self.assertEqual(receipt.line_items[0].confidence, "verified")

    def test_ingest_receipt_discards_unverified_pdf_line_items(self):
        parsed = Receipt("invoice", "invoice.pdf", date(2026, 6, 20), 17_020)
        boxes = [
            TextBox("排烟机", 10, 100), TextBox("件", 300, 100),
            TextBox("168.51", 430, 100),
        ]

        with patch("core.materials_assistant.extract_pdf_text", return_value="electronic invoice"), patch(
            "core.materials_assistant.extract_pdf_text_boxes", return_value=boxes
        ), patch("core.materials_assistant.parse_receipt_text", return_value=parsed), patch(
            "core.materials_assistant._recognize_locally", return_value=OCRResult(provider="paddle")
        ):
            receipt = ingest_receipt("invoice.pdf", provider_name="paddle")

        self.assertEqual(receipt.line_items, ())

    def test_mock_mode_discards_unverified_pdf_line_items(self):
        parsed = Receipt("invoice", "invoice.pdf", date(2026, 4, 22), 17_020)
        boxes = [
            TextBox("排烟机", 10, 100), TextBox("件", 300, 100),
            TextBox("168.51", 430, 100),
        ]

        with patch("core.materials_assistant.extract_pdf_text", return_value="electronic invoice"), patch(
            "core.materials_assistant.extract_pdf_text_boxes", return_value=boxes
        ), patch("core.materials_assistant.parse_receipt_text", return_value=parsed):
            receipt = ingest_receipt("invoice.pdf", provider_name="mock")

        self.assertEqual(receipt.line_items, ())

    def test_inventory_documents_expand_template_and_keep_pair_consistent(self):
        lines = [
            InventoryLine(date(2026, 4, 22), "排烟机", "排烟机", "个", "2", 8510, 17020, "临沂市国金护理用品有限公司"),
            InventoryLine(date(2026, 6, 20), "C620 无刷电机调速器", "C620", "件", "6", 16100, 96600, "深圳市睿炽科技有限公司"),
            InventoryLine(date(2026, 7, 11), "STM32 开发板", "F4", "块", "1", 3380, 3380, "深圳市某某电子科技有限公司"),
        ]

        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx", is_inbound=True)
            self._write_template(templates / "出库单_模板.xlsx", is_inbound=False)

            output_dir = root / "output"
            generated = render_inventory_documents(templates, output_dir, lines)

            self.assertEqual(set(generated), {"inbound", "outbound"})
            inbound = load_workbook(generated["inbound"], data_only=False).active
            outbound = load_workbook(generated["outbound"], data_only=False).active

            self.assertEqual(inbound["A5"].value.date(), date(2026, 4, 22))
            self.assertEqual(inbound["H5"].value, "临沂市国金护理用品有限公司")
            self.assertIsNone(outbound["H5"].value)
            self.assertEqual(inbound["G8"].value, "=SUM(G5:G7)")
            self.assertEqual(outbound["G8"].value, "=SUM(G5:G7)")
            self.assertIn("I8:J8", {str(item) for item in inbound.merged_cells.ranges})
            self.assertEqual(inbound.print_area, "'Sheet1'!$A$1:$J$9")

    def test_inventory_documents_store_user_text_as_text_not_formulas(self):
        line = InventoryLine(
            date(2026, 7, 11),
            '=HYPERLINK("https://example.invalid","材料")',
            "+SUM(A1:A2)",
            "-1+2",
            "1",
            10_000,
            10_000,
            "=1+1",
        )

        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx", is_inbound=True)
            self._write_template(templates / "出库单_模板.xlsx", is_inbound=False)

            generated = render_inventory_documents(templates, root / "output", [line])
            workbook = load_workbook(generated["inbound"], data_only=False)
            try:
                sheet = workbook.active
                for address, expected in {
                    "B5": line.name,
                    "C5": line.specification,
                    "D5": line.unit,
                    "H5": line.supplier_name,
                }.items():
                    self.assertEqual(sheet[address].value, expected)
                    self.assertEqual(sheet[address].data_type, "s")
            finally:
                workbook.close()

    def test_inventory_expansion_updates_all_total_cell_reference_styles(self):
        lines = [self._inventory_line(), self._inventory_line(), self._inventory_line()]
        formulas = {
            "relative": '=TEXT(G7,"[DBNUM2]")&"元整"',
            "absolute": '=TEXT($G$7,"[DBNUM2]")&"元整"',
            "absolute_row": '=TEXT(G$7,"[DBNUM2]")&"元整"',
            "absolute_column": '=TEXT($G7,"[DBNUM2]")&"元整"',
        }
        expected_formulas = {
            "relative": '=TEXT(G8,"[DBNUM2]")&"元整"',
            "absolute": '=TEXT($G$8,"[DBNUM2]")&"元整"',
            "absolute_row": '=TEXT(G$8,"[DBNUM2]")&"元整"',
            "absolute_column": '=TEXT($G8,"[DBNUM2]")&"元整"',
        }

        for reference_style, formula in formulas.items():
            with self.subTest(reference_style=reference_style), TemporaryDirectory() as temporary_dir:
                root = Path(temporary_dir)
                templates = root / "templates"
                templates.mkdir()
                self._write_template(
                    templates / "入库单_模板.xlsx", True, capitalized_total_formula=formula
                )
                self._write_template(
                    templates / "出库单_模板.xlsx", False, capitalized_total_formula=formula
                )

                generated = render_inventory_documents(templates, root / "output", lines)

                for document in generated.values():
                    sheet = load_workbook(document, data_only=False).active
                    self.assertEqual(sheet["I8"].value, expected_formulas[reference_style])

    def test_inventory_render_rejects_missing_template_before_creating_output(self):
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx", is_inbound=True)
            output_dir = root / "output"

            with self.assertRaisesRegex(FileNotFoundError, "出库单"):
                render_inventory_documents(templates, output_dir, [self._inventory_line()])

            self.assertFalse(output_dir.exists())

    def test_inventory_render_rejects_duplicate_template_before_creating_output(self):
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx", is_inbound=True)
            self._write_template(templates / "入库单_备用.xlsx", is_inbound=True)
            self._write_template(templates / "出库单_模板.xlsx", is_inbound=False)
            output_dir = root / "output"

            with self.assertRaisesRegex(ValueError, "入库单模板必须恰好一个"):
                render_inventory_documents(templates, output_dir, [self._inventory_line()])

            self.assertFalse(output_dir.exists())

    def test_inventory_render_rejects_broken_template_before_creating_output(self):
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            templates = root / "templates"
            templates.mkdir()
            self._write_template(templates / "入库单_模板.xlsx", is_inbound=True)
            (templates / "出库单_模板.xlsx").write_bytes(b"not a workbook")
            output_dir = root / "output"

            with self.assertRaisesRegex(ValueError, "出库单模板不可用"):
                render_inventory_documents(templates, output_dir, [self._inventory_line()])

            self.assertFalse(output_dir.exists())

    @staticmethod
    def _inventory_line():
        return InventoryLine(date(2026, 7, 11), "开发板", "F4", "块", "1", 10_000, 10_000)

    @staticmethod
    def _write_template(path, is_inbound, capitalized_total_formula='=TEXT(G7,"[DBNUM2]")&"元整"'):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.merge_cells("A2:J2")
        sheet["A2"] = "湘潭大学教学科研易耗品和材料" + ("入库单" if is_inbound else "出库单")
        sheet["A3"] = "单位(盖章)："
        sheet["F3"] = "经费代码："
        headers = [
            "入库日期" if is_inbound else "出库日期",
            "产品名称",
            "规格型号",
            "单位",
            "入库数量" if is_inbound else "出库数量",
            "单价(元)",
            "金额(元)",
            "供货单位" if is_inbound else "用途",
            "经办人" if is_inbound else "领用人",
            "管理员",
        ]
        for column, header in enumerate(headers, start=1):
            sheet.cell(4, column).value = header
        for row in (5, 6):
            for column in range(1, 11):
                sheet.cell(row, column).number_format = "0.00"
        sheet["F7"] = "合计："
        sheet["G7"] = "=SUM(G5:G6)"
        sheet["H7"] = "大写："
        sheet.merge_cells("I7:J7")
        sheet["I7"] = capitalized_total_formula
        sheet["A8"] = "注：模板说明"
        sheet.page_setup.orientation = "landscape"
        try:
            workbook.save(path)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
